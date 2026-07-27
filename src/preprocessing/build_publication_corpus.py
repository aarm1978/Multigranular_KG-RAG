"""Build the deterministic Phase A corpus for curated CIROH publications.

The curated Excel workbook defines corpus membership. BibTeX supplies ordered
bibliographic structures, declarative YAML resolves documented curation
exceptions, and Marker Markdown provides mechanically identifiable content.
Validation completes before any output file is written.
"""

from __future__ import annotations

import argparse
from collections import Counter
from dataclasses import dataclass
import hashlib
import json
import os
from pathlib import Path
import re
import sys
import tempfile
import unicodedata
from typing import Any, Iterable, Mapping, Sequence
from urllib.parse import urlsplit, urlunsplit

from openpyxl import load_workbook
from pybtex.database import Entry, Person
from pybtex.database.input.bibtex import Parser as BibtexParser
import yaml


SCHEMA_VERSION = "1.1.0"
PHASE_A_VERSION = "1.0.9"
OVERRIDE_SCHEMA_VERSION = "1.1.0"
PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_RAW_ROOT = Path("data/raw/papers")
DEFAULT_OUTPUT = Path("data/interim/papers/ciroh_publication_corpus.json")
DEFAULT_OVERRIDES = Path("data/curation/papers/publication_curation_overrides.yaml")
DEFAULT_EXPECTED_RECORD_COUNT = 228
EXCEL_FILE = "bib_entries_metadata.xlsx"
BIBTEX_FILE = "publications.bib"
REQUIRED_EXCEL_COLUMNS = ("id", "ZoteroID", "title", "year", "doi", "url", "journal")
ENTRY_TYPE_MAP = {
    "article": "journal_article",
    "inproceedings": "conference_paper",
    "inbook": "book_chapter",
}
KNOWN_OVERRIDE_ACTIONS = {"replace_bibliographic_record", "add_non_zotero_artifact"}
KNOWN_IDENTIFIER_DISPOSITION_ACTIONS = {"defer"}
KNOWN_IDENTIFIER_DISPOSITION_CONTEXTS = {"reference", "availability"}
REFERENCE_HEADINGS = {"references", "bibliography", "literature cited", "works cited"}
KEYWORD_HEADINGS = {"keywords", "key words", "author keywords", "index terms"}
AVAILABILITY_HEADINGS = {
    "data availability": "data_availability",
    "data availability statement": "data_availability",
    "data and code availability": "data_and_code_availability",
    "code and data availability": "data_and_code_availability",
    "code availability": "code_availability",
    "software availability": "software_availability",
    "availability of data and materials": "data_and_materials_availability",
}
TERMINAL_SECTION_HEADINGS = {
    "references",
    "bibliography",
    "literature cited",
    "works cited",
    "acknowledgments",
    "acknowledgements",
    "author contributions",
    "funding",
    "institutional review board statement",
    "informed consent statement",
    "conflicts of interest",
    "conflict of interest",
    "declarations",
    "supplementary materials",
    "appendix",
}
ABSTRACT_STOP_LABELS = {
    *KEYWORD_HEADINGS,
    "correspondence",
    "corresponding author",
    "author information",
    "author affiliation",
    "author affiliations",
    "affiliation",
    "affiliations",
}
STRUCTURED_ABSTRACT_LABELS = {
    "background",
    "context",
    "objective",
    "objectives",
    "purpose",
    "methods",
    "methodology",
    "results",
    "findings",
    "conclusions",
    "conclusion",
    "significance",
    "implications",
}
ABSTRACT_BODY_BOUNDARY_LABELS = {
    *ABSTRACT_STOP_LABELS,
    "introduction",
    "main text",
    "doi",
    "received",
    "revised",
    "accepted",
    "published",
    "manuscript received",
    "copyright",
    "license",
}
KEYWORD_METADATA_BOUNDARY_LABELS = {
    "abstract",
    "summary",
    "introduction",
    "doi",
    "corresponding author",
    "correspondence",
    "supplemental information",
    "supplementary information",
    "manuscript received",
    "received",
    "revised",
    "accepted",
    "published",
    "copyright",
    "license",
    "affiliation",
    "affiliations",
    "author affiliation",
    "author affiliations",
}
# Precision is preferable to guessing boundaries in Marker-collapsed keyword lists.
AMBIGUOUS_KEYWORD_CHARACTER_THRESHOLD = 80
AMBIGUOUS_KEYWORD_TOKEN_THRESHOLD = 6
KEYWORD_SENTENCE_TOKEN_THRESHOLD = 10
ABSTRACT_AUDIT_CHARACTER_THRESHOLD = 5_000
STRICT_DOI_PATTERN = re.compile(r"10\.\d{4,9}/[-._;()/:A-Z0-9]+", re.IGNORECASE)
DOI_START_PATTERN = re.compile(r"10\.\d{4,9}/", re.IGNORECASE)
DOI_RESOLVER_PREFIX_PATTERN = re.compile(r"https?://(?:dx\.)?doi\.org/(?:doi:)?/?", re.IGNORECASE)
PLAIN_URL_PATTERN = re.compile(r"https?://[^\s<>\"'`]+", re.IGNORECASE)
AUTOLINK_PATTERN = re.compile(r"<(https?://[^<>\s]+)>", re.IGNORECASE)
HEADING_PATTERN = re.compile(r"^(#{1,6})\s+(.+?)\s*$")
NUMBERING_PATTERN = re.compile(r"^\s*\d+(?:\.\d+)*(?:[.)])?\s+")
FENCE_PATTERN = re.compile(r"^\s*(```+|~~~+)")
CONTROL_PATTERN = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
EMAIL_PATTERN = re.compile(r"\b[\w.+-]+@[\w.-]+\.[A-Za-z]{2,}\b")
IMAGE_PATTERN = re.compile(r"!\[[^\]]*\]\([^)]*\)")
FIGURE_TABLE_CAPTION_PATTERN = re.compile(r"^\s*(?:\*{0,2})?(?:fig(?:ure)?|table)\s*[.:\d]", re.IGNORECASE)
EDITORIAL_METADATA_PATTERN = re.compile(
    r"^\s*(?:manuscript\s+)?(?:received|revised|accepted|published)\b|"
    r"^\s*(?:copyright|©|doi\s*:|license\b)",
    re.IGNORECASE,
)
AFFILIATION_PATTERN = re.compile(
    r"^\s*(?:\d+|[*_]{1,2}|<sup>[^<]+</sup>)?\s*"
    r"(?:department|institute|university|college|school|center|centre|laboratory|faculty)\b",
    re.IGNORECASE,
)
ORCID_PATTERN = re.compile(r"\borcid\b", re.IGNORECASE)


class CorpusBuildError(ValueError):
    """Raised when an authoritative input or output invariant is violated."""


@dataclass(frozen=True)
class ParsedBibtex:
    """Parsed BibTeX entries indexed by their exact original keys."""

    entries: dict[str, Entry]
    raw_author_values: dict[str, list[str]]


@dataclass(frozen=True)
class MarkdownLink:
    """One balanced standard Markdown link occurrence."""

    start: int
    end: int
    label: str
    destination: str


@dataclass(frozen=True)
class ParsedDoiCandidate:
    """One source-bounded DOI parse, including a rejected split disposition."""

    value: str | None
    end: int
    raw_value: str
    repaired: bool
    rejection_reason: str | None


def resolve_path(path: Path) -> Path:
    """Resolve a CLI path against the repository root without requiring it to exist."""

    return path if path.is_absolute() else PROJECT_ROOT / path


def normalize_line_endings(text: str) -> str:
    """Remove a UTF-8 BOM and normalize CRLF/CR line endings to LF."""

    return text.removeprefix("\ufeff").replace("\r\n", "\n").replace("\r", "\n")


def read_utf8(path: Path) -> str:
    """Read a required UTF-8 text file with the contract's permitted normalization."""

    try:
        return normalize_line_endings(path.read_text(encoding="utf-8-sig"))
    except FileNotFoundError as exc:
        raise CorpusBuildError(f"Required file is missing: {path}") from exc
    except UnicodeDecodeError as exc:
        raise CorpusBuildError(f"Required text is not valid UTF-8: {path}: {exc}") from exc


def sanitize_control_characters(text: str) -> str:
    """Replace each forbidden C0 character with one position-preserving space."""

    return CONTROL_PATTERN.sub(" ", text)


def audit_control_characters(text: str, source_path: str) -> dict[str, Any] | None:
    """Describe forbidden controls in an immutable raw Markdown source."""

    matches = list(CONTROL_PATTERN.finditer(text))
    if not matches:
        return None
    return {
        "category": "unexpected_control_characters",
        "detail": {
            "source_path": source_path,
            "occurrence_count": len(matches),
            "code_points": sorted({f"U+{ord(match.group(0)):04X}" for match in matches}),
            "line_numbers": sorted({text.count("\n", 0, match.start()) + 1 for match in matches}),
        },
    }


def sanitize_marker_value(value: Any) -> Any:
    """Recursively sanitize retained Marker-derived string values."""

    if isinstance(value, str):
        return sanitize_control_characters(value)
    if isinstance(value, list):
        return [sanitize_marker_value(item) for item in value]
    if isinstance(value, dict):
        return {key: sanitize_marker_value(item) for key, item in value.items()}
    return value


def find_forbidden_control_characters(value: Any, path: str = "$") -> list[dict[str, Any]]:
    """Find forbidden controls recursively in every emitted string value."""

    findings: list[dict[str, Any]] = []
    if isinstance(value, str):
        matches = list(CONTROL_PATTERN.finditer(value))
        if matches:
            findings.append(
                {
                    "path": path,
                    "occurrence_count": len(matches),
                    "code_points": sorted({f"U+{ord(match.group(0)):04X}" for match in matches}),
                }
            )
    elif isinstance(value, list):
        for index, item in enumerate(value):
            findings.extend(find_forbidden_control_characters(item, f"{path}[{index}]"))
    elif isinstance(value, dict):
        for key, item in value.items():
            findings.extend(find_forbidden_control_characters(item, f"{path}.{key}"))
    return findings


def clean_scalar(value: Any) -> Any:
    """Convert spreadsheet/YAML scalar absence to ``None`` and trim strings."""

    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    return value


def normalize_local_id(value: Any) -> str:
    """Normalize a technical local paper ID without changing suffix semantics."""

    if isinstance(value, bool) or value is None:
        raise CorpusBuildError(f"Invalid local paper ID: {value!r}")
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    local_id = str(value).strip()
    if not re.fullmatch(r"\d+(?:-[A-Za-z0-9][A-Za-z0-9_-]*)?", local_id):
        raise CorpusBuildError(f"Invalid local paper ID: {value!r}")
    return local_id


def natural_publication_key(local_id: str) -> tuple[int, int, str]:
    """Sort numeric IDs naturally and suffix artifacts immediately after their base."""

    match = re.fullmatch(r"(\d+)(?:-(.+))?", local_id)
    if not match:
        raise CorpusBuildError(f"Invalid local paper ID during sorting: {local_id!r}")
    suffix = match.group(2)
    return (int(match.group(1)), 0 if suffix is None else 1, suffix or "")


def stored_raw_root(raw_root: Path) -> str:
    """Return a deterministic non-absolute root label for persisted source paths."""

    resolved = raw_root.resolve()
    try:
        return resolved.relative_to(PROJECT_ROOT.resolve()).as_posix()
    except ValueError:
        return raw_root.name


def stored_path(raw_root: Path, path: Path) -> str:
    """Convert a raw file path to a deterministic root-relative persisted path."""

    try:
        relative = path.resolve().relative_to(raw_root.resolve())
    except ValueError as exc:
        raise CorpusBuildError(f"Source path escapes raw root: {path}") from exc
    return (Path(stored_raw_root(raw_root)) / relative).as_posix()


def normalize_doi(value: Any) -> str | None:
    """Normalize a DOI to its lowercase bare value."""

    text = clean_scalar(value)
    if text is None:
        return None
    text = str(text).strip()
    text = re.sub(r"(?i)^doi\s*:\s*", "", text)
    text = re.sub(r"(?i)^https?://(?:dx\.)?doi\.org/", "", text)
    text = text.strip().rstrip(".,;:")
    while text.endswith(")") and text.count("(") < text.count(")"):
        text = text[:-1]
    if not re.fullmatch(r"10\.\d{4,9}/\S+", text, re.IGNORECASE):
        return None
    return text.lower()


def normalize_url(value: Any) -> str | None:
    """Normalize an absolute HTTP(S) URL conservatively."""

    text = clean_scalar(value)
    if text is None:
        return None
    text = str(text).strip()
    try:
        split = urlsplit(text)
    except ValueError:
        return None
    if split.scheme.lower() not in {"http", "https"} or not split.netloc:
        return None
    return urlunsplit((split.scheme.lower(), split.netloc.lower(), split.path, split.query, split.fragment))


def make_identifier(scheme: str, value: str) -> dict[str, str]:
    """Create the fixed identifier representation."""

    uri = f"https://doi.org/{value}" if scheme == "doi" else value
    return {"scheme": scheme, "value": value, "uri": uri}


def comparison_text(value: Any) -> str:
    """Normalize harmless presentation differences for conflict comparison."""

    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value))
    text = text.replace("’", "'").replace("‘", "'")
    return " ".join(text.split()).casefold()


def normalize_heading_text(text: str) -> str:
    """Normalize a Markdown heading for controlled-section matching."""

    stripped = re.sub(r"\s+#+\s*$", "", text.strip())
    stripped = re.sub(r"<[^>]+>", " ", stripped)
    stripped = re.sub(r"[*_`]", "", stripped)
    stripped = NUMBERING_PATTERN.sub("", stripped)
    stripped = stripped.strip(" :.-")
    return " ".join(unicodedata.normalize("NFKC", stripped).casefold().split())


def load_excel_roster(path: Path) -> list[dict[str, Any]]:
    """Load the authoritative curated roster from its first worksheet."""

    if not path.is_file():
        raise CorpusBuildError(f"Excel roster is missing: {path}")
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # openpyxl exposes several format-specific exceptions
        raise CorpusBuildError(f"Excel roster cannot be read: {path}: {exc}") from exc
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration as exc:
        raise CorpusBuildError("Excel roster is empty") from exc
    headers = [str(value).strip() if value is not None else "" for value in header_row]
    missing = [column for column in REQUIRED_EXCEL_COLUMNS if column not in headers]
    if missing:
        raise CorpusBuildError(f"Excel roster missing required columns: {missing}")
    records: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row in rows:
        if not any(value is not None and str(value).strip() for value in row):
            continue
        mapped = {headers[index]: clean_scalar(value) for index, value in enumerate(row) if index < len(headers)}
        local_id = normalize_local_id(mapped.get("id"))
        if local_id in seen:
            raise CorpusBuildError(f"Duplicate Excel local paper ID: {local_id}")
        seen.add(local_id)
        mapped["id"] = local_id
        records.append(mapped)
    workbook.close()
    return records


def _extract_raw_author_values(text: str, original_keys: Sequence[str]) -> dict[str, list[str]]:
    """Capture deterministic raw author strings for output traceability."""

    result: dict[str, list[str]] = {}
    entry_pattern = re.compile(r"(?m)^\s*@[A-Za-z]+\s*[{(]\s*([^,\r\n]+),")
    matches = list(entry_pattern.finditer(text))
    for index, match in enumerate(matches):
        key = match.group(1).strip()
        body_end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        body = text[match.end() : body_end]
        author_match = re.search(r"(?i)\bauthor\s*=\s*", body)
        if not author_match:
            continue
        start = author_match.end()
        if start >= len(body) or body[start] not in {'{', '"'}:
            continue
        opener = body[start]
        if opener == '"':
            end = start + 1
            while end < len(body):
                if body[end] == '"' and body[end - 1] != "\\":
                    break
                end += 1
            raw_authors = body[start + 1 : end]
        else:
            depth = 1
            end = start + 1
            while end < len(body) and depth:
                if body[end] == "{" and body[end - 1] != "\\":
                    depth += 1
                elif body[end] == "}" and body[end - 1] != "\\":
                    depth -= 1
                end += 1
            raw_authors = body[start + 1 : end - 1] if depth == 0 else ""
        if raw_authors:
            parts = re.split(r"\s+and\s+", raw_authors)
            result[key] = [part.strip() for part in parts if part.strip()]
    return {key: result.get(key, []) for key in original_keys}


def load_bibtex(path: Path) -> ParsedBibtex:
    """Parse BibTeX with pybtex while preserving Zotero keys containing spaces."""

    text = read_utf8(path)
    header_pattern = re.compile(r"(?m)^(\s*@[A-Za-z]+\s*[{(])([^,\r\n]+)(,)")
    original_keys: list[str] = []

    def replace_key(match: re.Match[str]) -> str:
        original = match.group(2).strip()
        if original in original_keys:
            raise CorpusBuildError(f"Duplicate BibTeX key: {original}")
        original_keys.append(original)
        return f"{match.group(1)}ciroh_entry_{len(original_keys):06d}{match.group(3)}"

    transformed = header_pattern.sub(replace_key, text)
    transformed = '@string{june = "june"}\n@string{july = "july"}\n@string{sept = "sept"}\n' + transformed
    try:
        parsed = BibtexParser().parse_string(transformed)
    except Exception as exc:
        raise CorpusBuildError(f"BibTeX parsing failed: {path}: {exc}") from exc
    if len(parsed.entries) != len(original_keys):
        raise CorpusBuildError(
            f"BibTeX entry count mismatch after parsing: headers={len(original_keys)}, parsed={len(parsed.entries)}"
        )
    entries: dict[str, Entry] = {}
    for index, original_key in enumerate(original_keys, start=1):
        surrogate = f"ciroh_entry_{index:06d}"
        if surrogate not in parsed.entries:
            raise CorpusBuildError(f"BibTeX parser omitted entry: {original_key}")
        entries[original_key] = parsed.entries[surrogate]
    return ParsedBibtex(entries=entries, raw_author_values=_extract_raw_author_values(text, original_keys))


def load_overrides(path: Path) -> dict[str, dict[str, Any]]:
    """Load and validate the declarative curation override mapping."""

    if not path.is_file():
        raise CorpusBuildError(f"Override file is missing: {path}")
    try:
        value = yaml.safe_load(read_utf8(path))
    except yaml.YAMLError as exc:
        raise CorpusBuildError(f"Override YAML is invalid: {path}: {exc}") from exc
    if not isinstance(value, dict) or value.get("schema_version") != OVERRIDE_SCHEMA_VERSION:
        raise CorpusBuildError("Override file has an unsupported or missing schema_version")
    records = value.get("records")
    if not isinstance(records, dict):
        raise CorpusBuildError("Override file records must be a mapping")
    normalized: dict[str, dict[str, Any]] = {}
    for raw_id, override in records.items():
        local_id = normalize_local_id(raw_id)
        if not isinstance(override, dict):
            raise CorpusBuildError(f"Override record must be a mapping for {local_id}")
        action = override.get("action")
        dispositions = override.get("identifier_dispositions", [])
        if action is not None:
            if action not in KNOWN_OVERRIDE_ACTIONS:
                raise CorpusBuildError(f"Invalid override action for {local_id}")
            if not isinstance(override.get("metadata"), dict):
                raise CorpusBuildError(f"Override metadata must be a mapping for {local_id}")
        if not isinstance(dispositions, list):
            raise CorpusBuildError(f"Identifier dispositions must be a list for {local_id}")
        if action is None and not dispositions:
            raise CorpusBuildError(f"Override record {local_id} has no curation action")
        seen_dispositions: set[str] = set()
        for index, disposition in enumerate(dispositions, start=1):
            label = f"identifier disposition {index} for {local_id}"
            if not isinstance(disposition, dict):
                raise CorpusBuildError(f"{label} must be a mapping")
            required = {"context", "candidate", "action", "reason"}
            optional = {"occurrence"}
            if set(disposition) - required - optional or not required.issubset(disposition):
                raise CorpusBuildError(f"{label} has invalid or missing fields")
            if disposition["context"] not in KNOWN_IDENTIFIER_DISPOSITION_CONTEXTS:
                raise CorpusBuildError(f"{label} has an invalid context")
            if disposition["action"] not in KNOWN_IDENTIFIER_DISPOSITION_ACTIONS:
                raise CorpusBuildError(f"{label} has an invalid action")
            candidate = disposition["candidate"]
            if normalize_extracted_doi(candidate) != candidate:
                raise CorpusBuildError(f"{label} candidate is not an exact normalized DOI")
            if not clean_scalar(disposition["reason"]):
                raise CorpusBuildError(f"{label} lacks a reason")
            occurrence = disposition.get("occurrence")
            if occurrence is not None:
                allowed_occurrence_fields = {"evidence_text", "section", "line_start", "line_end"}
                if (
                    not isinstance(occurrence, dict)
                    or not occurrence
                    or set(occurrence) - allowed_occurrence_fields
                    or not clean_scalar(occurrence.get("evidence_text"))
                ):
                    raise CorpusBuildError(
                        f"{label} occurrence must include exact evidence_text and only supported fields"
                    )
            signature = json.dumps(disposition, ensure_ascii=False, sort_keys=True)
            if signature in seen_dispositions:
                raise CorpusBuildError(f"Duplicate {label}")
            seen_dispositions.add(signature)
        normalized[local_id] = override
    return normalized


def repair_zotero_key(original: str, bibtex_keys: Iterable[str]) -> str | None:
    """Return one exact BibTeX key after a deterministic reversible repair."""

    available = set(bibtex_keys)
    candidates: set[str] = set()
    for encoding in ("mac_roman", "latin-1", "cp1252"):
        try:
            repaired = original.encode(encoding).decode("utf-8")
        except (UnicodeEncodeError, UnicodeDecodeError):
            continue
        if repaired != original and repaired in available:
            candidates.add(repaired)
    if len(candidates) > 1:
        raise CorpusBuildError(f"Zotero key repair is ambiguous for {original!r}: {sorted(candidates)}")
    return next(iter(candidates), None)


def person_record(person: Person, position: int, raw_value: str | None = None, display: str | None = None) -> dict[str, Any]:
    """Convert a pybtex person to the fixed ordered-author representation."""

    given = [str(value) for value in (*person.first_names, *person.middle_names)]
    particles = [str(value) for value in person.prelast_names]
    family_parts = [str(value) for value in person.last_names]
    suffix_parts = [str(value) for value in person.lineage_names]
    family = " ".join(family_parts) or None
    suffix = " ".join(suffix_parts) or None
    natural = " ".join([*given, *particles, *family_parts, *suffix_parts]).strip()
    literal = None if given else (family if family else str(person))
    return {
        "position": position,
        "display_name": display or natural or str(person),
        "given_names": given,
        "family_name": family,
        "name_particles": particles,
        "suffix": suffix,
        "literal_name": literal,
        "raw_bibtex": raw_value or str(person),
    }


def authors_from_entry(entry: Entry, raw_values: Sequence[str]) -> list[dict[str, Any]]:
    """Preserve BibTeX author order and structure."""

    people = entry.persons.get("author", [])
    return [
        person_record(person, index, raw_values[index - 1] if index <= len(raw_values) else None)
        for index, person in enumerate(people, start=1)
    ]


def authors_from_override(values: Any) -> list[dict[str, Any]]:
    """Interpret ordered authors from a declarative override."""

    if not isinstance(values, list):
        raise CorpusBuildError("Override authors must be a list")
    authors: list[dict[str, Any]] = []
    for index, value in enumerate(values, start=1):
        if not isinstance(value, dict) or not clean_scalar(value.get("raw_name")):
            raise CorpusBuildError(f"Override author {index} lacks raw_name")
        raw_name = str(value["raw_name"]).strip()
        authors.append(person_record(Person(raw_name), index, raw_name, clean_scalar(value.get("display_name"))))
    return authors


def extract_headings(markdown: str) -> list[dict[str, Any]]:
    """Extract ATX headings outside fenced code blocks in source order."""

    headings: list[dict[str, Any]] = []
    active_fence: str | None = None
    for line_number, line in enumerate(markdown.splitlines(), start=1):
        fence = FENCE_PATTERN.match(line)
        if fence:
            marker = fence.group(1)[0]
            active_fence = None if active_fence == marker else (marker if active_fence is None else active_fence)
            continue
        if active_fence is not None:
            continue
        match = HEADING_PATTERN.match(line)
        if not match:
            continue
        text = re.sub(r"\s+#+\s*$", "", match.group(2)).strip()
        headings.append(
            {
                "level": len(match.group(1)),
                "text": text,
                "normalized_text": normalize_heading_text(text),
                "line_number": line_number,
            }
        )
    return headings


def section_bounds(headings: Sequence[Mapping[str, Any]], index: int, line_count: int) -> tuple[int, int]:
    """Return bounds ending immediately before the next heading of any level."""

    heading = headings[index]
    start = int(heading["line_number"]) + 1
    end = int(headings[index + 1]["line_number"]) - 1 if index + 1 < len(headings) else line_count
    return start, max(start - 1, end)


def reference_section_bounds(
    headings: Sequence[Mapping[str, Any]], index: int, line_count: int
) -> tuple[int, int]:
    """Return reference bounds ending only at a clear subsequent terminal heading."""

    start = int(headings[index]["line_number"]) + 1
    end = line_count
    for later in headings[index + 1 :]:
        normalized = str(later["normalized_text"])
        if normalized in TERMINAL_SECTION_HEADINGS and normalized not in REFERENCE_HEADINGS:
            end = int(later["line_number"]) - 1
            break
    return start, max(start - 1, end)


def clean_section_text(lines: Sequence[str], start: int, end: int) -> str | None:
    """Join a one-based line range without rewriting source content."""

    if end < start:
        return None
    value = "\n".join(lines[start - 1 : end]).strip()
    return value or None


def _strip_line_presentation(line: str) -> str:
    """Remove mechanical line-level presentation for controlled-label matching."""

    value = re.sub(r"^\s*#{1,6}\s+", "", line.strip())
    value = re.sub(r"<[^>]+>", " ", value)
    value = re.sub(r"^\s*(?:[-+•]|\d+[.)])\s+", "", value)
    value = value.strip()
    while value.startswith(("**", "__")):
        value = value[2:].lstrip()
    while value.startswith(("*", "_", "`")):
        value = value[1:].lstrip()
    return value


def match_controlled_label(line: str, labels: set[str]) -> tuple[str, str] | None:
    """Match an exact or colon-delimited controlled label with Markdown emphasis."""

    value = _strip_line_presentation(line)
    for label in sorted(labels, key=lambda item: (-len(item), item)):
        match = re.match(rf"(?i)^{re.escape(label)}\b", value)
        if not match:
            continue
        remainder = value[match.end() :].lstrip()
        emphasized = bool(re.match(r"^\s*(?:\*\*|__|\*|_)", line))
        if remainder.startswith(":"):
            remainder = remainder[1:].strip()
            remainder = re.sub(r"^(?:\*\*|__|\*|_|`)+", "", remainder).lstrip()
            return label, remainder
        remainder = re.sub(r"^(?:\*\*|__|\*|_|`)+", "", remainder).lstrip()
        if not remainder or emphasized:
            return label, remainder.strip()
    return None


def is_abstract_boundary(line: str) -> bool:
    """Return whether a line clearly begins post-abstract metadata."""

    if match_controlled_label(line, ABSTRACT_STOP_LABELS):
        return True
    if re.search(r"\b[\w.+-]+@[\w.-]+\.[A-Za-z]{2,}\b", line):
        return True
    normalized = normalize_heading_text(_strip_line_presentation(line))
    if normalized in ABSTRACT_STOP_LABELS:
        return True
    return bool(
        re.match(
            r"^\s*(?:\d+|[*_]{1,2})?\s*(?:department|institute|university|college|school|center|centre)\b",
            _strip_line_presentation(line),
            re.IGNORECASE,
        )
    )


def controlled_label_offset(line: str, labels: set[str]) -> int | None:
    """Locate a controlled label at line start or in explicit emphasis markup."""

    if match_controlled_label(line, labels):
        return 0
    offsets: list[int] = []
    for label in labels:
        match = re.search(rf"(?:\*\*|__)\s*{re.escape(label)}\s*:?(?:\*\*|__)?", line, re.IGNORECASE)
        if match:
            offsets.append(match.start())
    return min(offsets) if offsets else None


def abstract_boundary_offset(line: str) -> int | None:
    """Locate the earliest post-abstract metadata boundary within a line."""

    offsets: list[int] = []
    label_offset = controlled_label_offset(line, ABSTRACT_STOP_LABELS)
    if label_offset is not None:
        offsets.append(label_offset)
    email = re.search(r"\b[\w.+-]+@[\w.-]+\.[A-Za-z]{2,}\b", line)
    if email:
        offsets.append(email.start())
    if is_abstract_boundary(line):
        offsets.append(0)
    return min(offsets) if offsets else None


def is_terminal_section_line(line: str) -> bool:
    """Return whether a non-heading line clearly begins a terminal section."""

    return controlled_label_offset(line, TERMINAL_SECTION_HEADINGS) == 0


def terminal_section_offset(line: str) -> int | None:
    """Locate a terminal-section label, including an inline bold label."""

    return controlled_label_offset(line, TERMINAL_SECTION_HEADINGS)


def trim_blank_line_bounds(lines: Sequence[str], start: int, end: int) -> tuple[int, int]:
    """Trim blank lines from one-based inclusive source bounds."""

    while start <= end and not lines[start - 1].strip():
        start += 1
    while end >= start and not lines[end - 1].strip():
        end -= 1
    return start, end


def make_extraction_warning(
    category: str,
    reason: str,
    line_number: int | None = None,
    candidate: str | None = None,
) -> dict[str, Any]:
    """Build a deterministic structured warning for content extraction."""

    detail: dict[str, Any] = {"reason": reason, "line_number": line_number}
    if candidate is not None:
        detail["candidate"] = candidate
    return {"category": category, "detail": detail}


def _line_boundary_reason(line: str, labels: set[str]) -> str | None:
    """Classify a deterministic metadata or structural boundary line."""

    if IMAGE_PATTERN.search(line):
        return "contains_image"
    if FIGURE_TABLE_CAPTION_PATTERN.match(_strip_line_presentation(line)):
        return "contains_figure_or_table_caption"
    if EMAIL_PATTERN.search(line):
        return "contains_correspondence"
    if AFFILIATION_PATTERN.match(_strip_line_presentation(line)):
        return "contains_affiliation"
    matched = match_controlled_label(line, labels)
    if matched:
        label = matched[0]
        if label in {"introduction", "main text", "abstract", "summary"}:
            return "contains_body_section"
        if label in KEYWORD_HEADINGS:
            return "contains_keyword_declaration"
        return "contains_editorial_metadata"
    if EDITORIAL_METADATA_PATTERN.match(_strip_line_presentation(line)):
        return "contains_editorial_metadata"
    return None


def _starts_structured_abstract_block(line: str) -> bool:
    """Return whether a paragraph starts with an approved structured label."""

    return match_controlled_label(line, STRUCTURED_ABSTRACT_LABELS) is not None


def _strip_marker_leading_line_number(value: str) -> str:
    """Remove Marker line ordinals only when a dense sequential run confirms them."""

    match = re.match(r"^(\d{1,3})\s+(?=[A-Z])", value)
    if not match:
        return value
    first = int(match.group(1))
    number_matches = list(re.finditer(r"(?<![\w.])(\d{1,3})\s+(?=\S)", value))
    expected = first
    sequence: list[re.Match[str]] = []
    for number_match in number_matches:
        if int(number_match.group(1)) == expected:
            sequence.append(number_match)
            expected += 1
    if len(sequence) < 5 or sequence[0].start() != 0:
        return value
    cleaned = value
    for number_match in reversed(sequence):
        cleaned = cleaned[: number_match.start()] + cleaned[number_match.end() :]
    return cleaned


def validate_abstract_candidate(value: str) -> str | None:
    """Return a rejection reason when a Markdown abstract is structurally contaminated."""

    if IMAGE_PATTERN.search(value):
        return "contains_image"
    if any(FIGURE_TABLE_CAPTION_PATTERN.match(_strip_line_presentation(line)) for line in value.splitlines()):
        return "contains_figure_or_table_caption"
    if any(HEADING_PATTERN.match(line) for line in value.splitlines()):
        return "contains_body_section"
    if EMAIL_PATTERN.search(value) or any(
        AFFILIATION_PATTERN.match(_strip_line_presentation(line)) for line in value.splitlines()
    ):
        return "contains_correspondence_or_affiliation"
    if any(
        EDITORIAL_METADATA_PATTERN.match(_strip_line_presentation(line))
        for line in value.splitlines()
    ):
        return "contains_editorial_metadata"
    blocks = [block for block in re.split(r"\n\s*\n", value) if block.strip()]
    for block_index, block in enumerate(blocks[1:], start=1):
        previous_lines = blocks[block_index - 1].splitlines()
        previous_label = match_controlled_label(previous_lines[0], STRUCTURED_ABSTRACT_LABELS)
        follows_label_only = bool(previous_label and not previous_label[1] and len(previous_lines) == 1)
        if not follows_label_only and not _starts_structured_abstract_block(block.splitlines()[0]):
            return "crossed_paragraph_boundary"
    if len(value) > ABSTRACT_AUDIT_CHARACTER_THRESHOLD and len(blocks) > 1:
        return "excessive_contaminated_block"
    return None


def extract_abstract_with_disposition(
    markdown: str,
    headings: Sequence[Mapping[str, Any]],
    canonical_artifact: str,
) -> tuple[str | None, dict[str, Any] | None, str | None]:
    """Extract one logical Markdown abstract block and its rejection disposition."""

    lines = markdown.splitlines()
    for index, heading in enumerate(headings):
        if heading["normalized_text"] != "abstract":
            continue
        start, end = section_bounds(headings, index, len(lines))
        while start <= end and not lines[start - 1].strip():
            start += 1
        if start > end:
            return None, None, None

        blocks: list[tuple[int, int, list[str]]] = []
        cursor = start
        rejection_reason: str | None = None
        while cursor <= end:
            block_start = cursor
            block_lines: list[str] = []
            while cursor <= end and lines[cursor - 1].strip():
                line = lines[cursor - 1]
                boundary_offset = controlled_label_offset(line, ABSTRACT_BODY_BOUNDARY_LABELS)
                if boundary_offset is not None and boundary_offset > 0:
                    prefix = line[:boundary_offset].rstrip()
                    if prefix:
                        block_lines.append(prefix)
                    cursor += 1
                    break
                reason = _line_boundary_reason(line, ABSTRACT_BODY_BOUNDARY_LABELS)
                if reason:
                    rejection_reason = reason if not block_lines and not blocks else None
                    break
                block_lines.append(line)
                cursor += 1
            if block_lines:
                blocks.append((block_start, cursor - 1, block_lines))
            if rejection_reason or cursor > end or (cursor <= end and lines[cursor - 1].strip()):
                break
            while cursor <= end and not lines[cursor - 1].strip():
                cursor += 1
            if cursor > end:
                break
            current_label = match_controlled_label(block_lines[0], STRUCTURED_ABSTRACT_LABELS)
            current_is_label_only = bool(
                current_label and not current_label[1] and len(block_lines) == 1
            )
            if not current_is_label_only and not _starts_structured_abstract_block(lines[cursor - 1]):
                break

        if not blocks:
            return None, None, rejection_reason
        abstract = "\n\n".join("\n".join(block[2]).strip() for block in blocks).strip()
        abstract = _strip_marker_leading_line_number(abstract)
        rejection_reason = validate_abstract_candidate(abstract)
        if rejection_reason:
            return None, None, rejection_reason
        return abstract, {
                "source_type": "markdown_explicit",
                "source_artifact": canonical_artifact,
                "section": heading["text"],
                "line_start": blocks[0][0],
                "line_end": blocks[-1][1],
            }, None
    return None, None, None


def extract_abstract(
    markdown: str,
    headings: Sequence[Mapping[str, Any]],
    canonical_artifact: str,
) -> tuple[str | None, dict[str, Any] | None]:
    """Extract an explicitly headed logical Markdown abstract, if valid."""

    abstract, source, _ = extract_abstract_with_disposition(markdown, headings, canonical_artifact)
    return abstract, source


def normalize_keyword_separators(value: str) -> str:
    """Normalize explicit typographic and TeX keyword separators."""

    normalized = unicodedata.normalize("NFKC", value)
    normalized = re.sub(r"\\\(\s*\\(?:cdot|bullet)\s*\\\)", " · ", normalized, flags=re.IGNORECASE)
    normalized = re.sub(r"\$\s*\\(?:cdot|bullet)\s*\$", " · ", normalized, flags=re.IGNORECASE)
    normalized = re.sub(r"\\(?:cdot|bullet)\b", " · ", normalized, flags=re.IGNORECASE)
    return re.sub(r"[|｜¦ǀ]", " | ", normalized)


def has_explicit_keyword_separator(value: str) -> bool:
    """Return whether source text provides a mechanical keyword delimiter."""

    return bool(re.search(r"[,;·•|\n]", normalize_keyword_separators(value)))


def split_keywords(value: str) -> list[str]:
    """Split only explicitly delimited keyword declarations."""

    return [part for part in re.split(r"[,;·•|\n]", normalize_keyword_separators(value)) if part.strip()]


def clean_keyword_value(value: str) -> str | None:
    """Clean presentation syntax from one explicit keyword without inferring content."""

    cleaned = unicodedata.normalize("NFKC", value).strip()
    cleaned = re.sub(r"^\s*(?:(?:[-+•]|\d+[.)])\s+|[—–]\s*)", "", cleaned)
    cleaned = re.sub(r"^(?:\*\*|__|\*|_|`)+\s*", "", cleaned)
    cleaned = re.sub(r"\s*(?:\*\*|__|\*|_|`)+$", "", cleaned)
    cleaned = cleaned.strip().rstrip(",;·•|¦ǀ").strip()
    if cleaned.endswith(")") and cleaned.count(")") > cleaned.count("("):
        cleaned = cleaned[:-1].rstrip()
    if cleaned.startswith("(") and cleaned.count("(") > cleaned.count(")"):
        cleaned = cleaned[1:].lstrip()
    if cleaned.endswith(".") and not re.fullmatch(r"(?:[A-Za-z]\.){2,}", cleaned):
        cleaned = cleaned[:-1].rstrip()
    if not cleaned or cleaned.startswith(("*", "_", "`")) or cleaned.endswith(("*", "_", "`")):
        return None
    if cleaned.count("`") % 2 or cleaned.count("**") % 2 or cleaned.count("__") % 2:
        return None
    return " ".join(cleaned.split())


def is_ambiguous_keyword_declaration(value: str) -> bool:
    """Identify long undelimited declarations without inferring semantic boundaries."""

    cleaned = clean_keyword_value(value)
    if not cleaned or has_explicit_keyword_separator(value):
        return False
    return (
        len(cleaned) > AMBIGUOUS_KEYWORD_CHARACTER_THRESHOLD
        or len(cleaned.split()) >= AMBIGUOUS_KEYWORD_TOKEN_THRESHOLD
    )


def validate_keyword_candidate(value: str) -> str | None:
    """Return a deterministic reason for rejecting front-matter contamination."""

    stripped = _strip_line_presentation(value)
    normalized = normalize_heading_text(stripped)
    if EMAIL_PATTERN.search(value):
        return "contains_email"
    if ORCID_PATTERN.search(value):
        return "contains_orcid"
    if re.search(r"https?://|\bdoi\s*:|10\.\d{4,9}/", value, re.IGNORECASE):
        return "contains_identifier_declaration"
    if EDITORIAL_METADATA_PATTERN.match(stripped) or re.search(
        r"\b(?:manuscript\s+)?(?:received|revised|accepted|published)\b|©|\bcopyright\b|\blicen[cs]ed?\b",
        value,
        re.IGNORECASE,
    ):
        return "contains_publication_metadata"
    if match_controlled_label(stripped, KEYWORD_METADATA_BOUNDARY_LABELS):
        return "contains_front_matter_label"
    if re.search(
        r"\b(?:department|university|college|school|institute|center|centre|laboratory|faculty)\b",
        value,
        re.IGNORECASE,
    ):
        return "contains_affiliation"
    if re.search(r"\bsupplement(?:al|ary)\s+(?:information|material)\b", value, re.IGNORECASE):
        return "contains_supplemental_information"
    if normalized in KEYWORD_METADATA_BOUNDARY_LABELS:
        return "contains_front_matter_label"
    if len(value.split()) >= KEYWORD_SENTENCE_TOKEN_THRESHOLD and re.search(r"[.!?]$", value.strip()):
        return "contains_sentence"
    return None


def keyword_boundary_reason(line: str) -> str | None:
    """Classify a line that terminates an explicit keyword declaration block."""

    reason = _line_boundary_reason(line, KEYWORD_METADATA_BOUNDARY_LABELS)
    if reason:
        return reason
    if ORCID_PATTERN.search(line):
        return "contains_orcid"
    return None


def collect_keyword_block(
    lines: Sequence[str],
    start: int,
    end: int,
) -> tuple[list[tuple[int, str]], dict[str, Any] | None]:
    """Collect the first contiguous keyword declaration block within source bounds."""

    while start <= end and not lines[start - 1].strip():
        start += 1
    collected: list[tuple[int, str]] = []
    cursor = start
    while cursor <= end and lines[cursor - 1].strip():
        reason = keyword_boundary_reason(lines[cursor - 1])
        if reason:
            warning = make_extraction_warning(
                "keyword_section_stopped_at_metadata",
                reason,
                cursor,
                lines[cursor - 1].strip(),
            )
            return collected, warning
        collected.append((cursor, lines[cursor - 1]))
        cursor += 1
    return collected, None


def keyword_candidates_from_declaration(
    value: str,
    line_number: int | None,
    source_type: str,
    canonical_artifact: str,
    section: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Normalize one explicit declaration without guessing missing separators."""

    if is_ambiguous_keyword_declaration(value):
        return [], [
            make_extraction_warning(
                "ambiguous_keyword_declaration",
                "undelimited_candidate_exceeds_precision_threshold",
                line_number,
                clean_keyword_value(value) or value.strip(),
            )
        ]
    candidates: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []
    for raw_part in split_keywords(value):
        raw_value = clean_keyword_value(raw_part)
        if not raw_value:
            continue
        reason = validate_keyword_candidate(raw_value)
        if reason:
            warnings.append(
                make_extraction_warning("rejected_keyword_candidate", reason, line_number, raw_value)
            )
            continue
        if is_ambiguous_keyword_declaration(raw_part):
            warnings.append(
                make_extraction_warning(
                    "ambiguous_keyword_declaration",
                    "undelimited_candidate_exceeds_precision_threshold",
                    line_number,
                    raw_value,
                )
            )
            continue
        candidates.append(
            {
                "value": raw_value.casefold(),
                "raw_value": raw_value,
                "source_type": source_type,
                "source_location": {
                    "source_artifact": canonical_artifact,
                    "section": section,
                    "line_start": line_number,
                    "line_end": line_number,
                },
            }
        )
    return candidates, warnings


def extract_keywords(
    markdown: str,
    headings: Sequence[Mapping[str, Any]],
    canonical_artifact: str,
    bibtex_keywords: str | None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Extract explicit Markdown and BibTeX keyword declarations."""

    lines = markdown.splitlines()
    candidates: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []
    occupied: set[int] = set()
    for index, heading in enumerate(headings):
        if heading["normalized_text"] not in KEYWORD_HEADINGS:
            continue
        start, end = section_bounds(headings, index, len(lines))
        occupied.add(int(heading["line_number"]))
        block, boundary_warning = collect_keyword_block(lines, start, end)
        occupied.update(line_number for line_number, _ in block)
        if boundary_warning:
            warnings.append(boundary_warning)
        found = False
        section_warning_start = len(warnings)
        for line_number, line in block:
            line_candidates, line_warnings = keyword_candidates_from_declaration(
                line,
                line_number,
                "markdown_explicit",
                canonical_artifact,
                str(heading["text"]),
            )
            found = found or bool(line_candidates)
            candidates.extend(line_candidates)
            warnings.extend(line_warnings)
        if not found and not any(
            item["category"] == "ambiguous_keyword_declaration"
            for item in warnings[section_warning_start:]
        ):
            warnings.append(
                make_extraction_warning(
                    "ambiguous_keyword_format",
                    "no_valid_keyword_candidate",
                    int(heading["line_number"]),
                    str(heading["text"]),
                )
            )
    for line_number, line in enumerate(lines, start=1):
        if line_number in occupied:
            continue
        matched = match_controlled_label(line, KEYWORD_HEADINGS)
        if not matched:
            continue
        label, remainder = matched
        inline_candidates, inline_warnings = keyword_candidates_from_declaration(
            remainder,
            line_number,
            "markdown_explicit",
            canonical_artifact,
            label,
        )
        candidates.extend(inline_candidates)
        warnings.extend(inline_warnings)
        if not inline_candidates and not inline_warnings:
            warnings.append(
                make_extraction_warning(
                    "ambiguous_keyword_format",
                    "empty_inline_keyword_declaration",
                    line_number,
                    line.strip(),
                )
            )
    if bibtex_keywords:
        bib_candidates, bib_warnings = keyword_candidates_from_declaration(
            bibtex_keywords,
            None,
            "bibtex_explicit",
            canonical_artifact,
            "keywords",
        )
        candidates.extend(bib_candidates)
        warnings.extend(bib_warnings)
    result: list[dict[str, Any]] = []
    seen: set[str] = set()
    candidates.sort(
        key=lambda item: (
            item["source_location"]["line_start"] is None,
            item["source_location"]["line_start"] or 0,
            item["source_location"]["line_end"] or 0,
        )
    )
    for candidate in candidates:
        if candidate["value"] and candidate["value"] not in seen:
            seen.add(candidate["value"])
            result.append(candidate)
    return result, warnings


def unescape_markdown_punctuation(value: str) -> str:
    """Remove Markdown escaping only for punctuation characters."""

    return re.sub(r"\\([!\"#$%&'()*+,\-./:;<=>?@\[\\\]^_`{|}~])", r"\1", value)


def scan_markdown_links(text: str) -> list[MarkdownLink]:
    """Scan balanced standard Markdown links, including parentheses in destinations."""

    links: list[MarkdownLink] = []
    index = 0
    while index < len(text):
        if text[index] != "[" or (index > 0 and text[index - 1] == "!"):
            index += 1
            continue
        label_start = index
        square_depth = 1
        cursor = index + 1
        while cursor < len(text) and square_depth:
            if text[cursor] == "[" and text[cursor - 1] != "\\":
                square_depth += 1
            elif text[cursor] == "]" and text[cursor - 1] != "\\":
                square_depth -= 1
            cursor += 1
        if square_depth:
            index += 1
            continue
        label_end = cursor - 1
        while cursor < len(text) and text[cursor].isspace():
            cursor += 1
        if cursor >= len(text) or text[cursor] != "(":
            index += 1
            continue
        outer_start = cursor
        cursor += 1
        while cursor < len(text) and text[cursor].isspace():
            cursor += 1
        destination_start = cursor
        if cursor < len(text) and text[cursor] == "<":
            destination_start = cursor + 1
            close = text.find(">", destination_start)
            if close < 0:
                index += 1
                continue
            destination_end = close
            cursor = close + 1
            while cursor < len(text) and text[cursor] != ")":
                cursor += 1
        else:
            nested_parentheses = 0
            while cursor < len(text):
                character = text[cursor]
                escaped = cursor > destination_start and text[cursor - 1] == "\\"
                if character == "(" and not escaped:
                    nested_parentheses += 1
                elif character == ")" and not escaped:
                    if nested_parentheses == 0:
                        break
                    nested_parentheses -= 1
                cursor += 1
            destination_end = cursor
        if cursor >= len(text) or text[cursor] != ")":
            index += 1
            continue
        destination = text[destination_start:destination_end].strip()
        if destination:
            links.append(
                MarkdownLink(
                    start=label_start,
                    end=cursor + 1,
                    label=text[label_start + 1 : label_end],
                    destination=destination,
                )
            )
        index = cursor + 1 if cursor + 1 > outer_start else index + 1
    return links


def mask_spans(text: str, spans: Iterable[tuple[int, int]]) -> str:
    """Mask character spans while preserving line length and positions."""

    characters = list(text)
    for start, end in spans:
        characters[start:end] = " " * (end - start)
    return "".join(characters)


def normalize_extracted_doi(value: Any) -> str | None:
    """Normalize and strictly validate a DOI extracted from publication text."""

    text = clean_scalar(value)
    if text is None:
        return None
    text = unescape_markdown_punctuation(str(text).strip())
    text = re.sub(r"(?i)^doi\s*:\s*", "", text)
    text = re.sub(r"(?i)^https?://(?:dx\.)?doi\.org/", "", text)
    while text and text[-1] in ".,;:!?":
        text = text[:-1]
    while text.endswith(")") and text.count(")") > text.count("("):
        text = text[:-1]
    if not text or text.count("(") != text.count(")"):
        return None
    if re.search(r"(?i)https?://|doi\.org|\]\(|[\[\]{}<>`\"']|\s", text):
        return None
    if not STRICT_DOI_PATTERN.fullmatch(text):
        return None
    if not re.search(r"[A-Z0-9)]$", text, re.IGNORECASE):
        return None
    return text.lower()


def _doi_continuation_has_split_evidence(current: str, token: str) -> bool:
    """Return whether adjacent formatting whitespace demonstrably split a DOI token."""

    token_without_prose_punctuation = token.rstrip(".,;:!?")
    if not token_without_prose_punctuation or not re.search(r"\d", token_without_prose_punctuation):
        return False
    lowered = token_without_prose_punctuation.casefold()
    if lowered.startswith(("http://", "https://", "doi.org", "www.", "10.")):
        return False
    if token_without_prose_punctuation.isdigit() and len(token_without_prose_punctuation) == 4:
        year = int(token_without_prose_punctuation)
        if 1800 <= year <= 2100:
            return False
    suffix = current.split("/", 1)[1] if "/" in current else ""
    return bool(
        current.endswith((".", ":", "/", "-"))
        or not re.search(r"\d", suffix)
        or token_without_prose_punctuation.startswith((".", ":", "/", "("))
        or token_without_prose_punctuation.isdigit()
        or (current.endswith(")") and re.search(r"[:()]", token_without_prose_punctuation))
    )


def parse_source_doi_candidate(text: str, start: int) -> ParsedDoiCandidate:
    """Parse one DOI and repair only source-adjacent formatting splits."""

    initial = STRICT_DOI_PATTERN.match(text, start)
    if initial is None:
        return ParsedDoiCandidate(None, start, "", False, "missing_doi_suffix")
    raw_value = initial.group(0)
    cursor = initial.end()
    repaired = False
    while cursor < len(text):
        whitespace = re.match(r"[ \t]*(?:\n[ \t]*)?", text[cursor:])
        if whitespace is None or not whitespace.group(0):
            break
        separator = whitespace.group(0)
        if "\n\n" in separator:
            break
        token_start = cursor + len(separator)
        token_match = re.match(r"\S+", text[token_start:])
        if token_match is None:
            break
        token = token_match.group(0)
        if not _doi_continuation_has_split_evidence(raw_value, token):
            break
        candidate = raw_value + token
        normalized = normalize_extracted_doi(candidate)
        if normalized is None:
            return ParsedDoiCandidate(
                None,
                token_start + len(token),
                candidate,
                repaired,
                "invalid_split_doi_continuation",
            )
        raw_value = candidate
        cursor = token_start + len(token)
        repaired = True
    if cursor < len(text) and not text[cursor].isspace() and text[cursor] in "%<>":
        return ParsedDoiCandidate(
            None,
            cursor + 1,
            raw_value + text[cursor],
            repaired,
            "invalid_doi_character_after_candidate",
        )
    normalized = normalize_extracted_doi(raw_value)
    return ParsedDoiCandidate(
        normalized,
        cursor,
        raw_value,
        repaired,
        None if normalized else "invalid_doi_candidate",
    )


def normalize_extracted_url(value: Any) -> str | None:
    """Normalize a text-derived absolute URL and reject Markdown contamination."""

    text = clean_scalar(value)
    if text is None:
        return None
    text = unescape_markdown_punctuation(str(text).strip())
    while text and text[-1] in ".,;:!?":
        text = text[:-1]
    while text.endswith(")") and text.count(")") > text.count("("):
        text = text[:-1]
    if re.search(r"\]\(|[\[\]{}<>`]", text) or re.search(r"(?i)https?://", text[8:]):
        return None
    normalized = normalize_url(text)
    if normalized is None or any(character.isspace() for character in normalized):
        return None
    return normalized


def doi_from_resolver_url(url: str) -> str | None:
    """Return a strict DOI when a URL is an exact DOI resolver URL."""

    try:
        split = urlsplit(url)
    except ValueError:
        return None
    if split.netloc.casefold() not in {"doi.org", "dx.doi.org", "www.doi.org"}:
        return None
    return normalize_extracted_doi(split.path.lstrip("/"))


def is_doi_resolver_url(url: str) -> bool:
    """Return whether an absolute URL targets a DOI resolver host."""

    try:
        return urlsplit(url).netloc.casefold() in {"doi.org", "dx.doi.org", "www.doi.org"}
    except ValueError:
        return False


def parse_doi_destination(value: str) -> ParsedDoiCandidate | None:
    """Parse a DOI-bearing Markdown destination with split diagnostics."""

    unescaped = unescape_markdown_punctuation(value.strip())
    resolver = DOI_RESOLVER_PREFIX_PATTERN.match(unescaped)
    if resolver:
        return parse_source_doi_candidate(unescaped, resolver.end())
    return parse_source_doi_candidate(unescaped, 0) if DOI_START_PATTERN.match(unescaped) else None


def classify_text_destination(value: str) -> tuple[str, str] | None:
    """Classify one Markdown destination or autolink as a strict DOI or URL."""

    unescaped = unescape_markdown_punctuation(value.strip())
    parsed = parse_doi_destination(value)
    if parsed and parsed.value:
        return ("doi", parsed.value)
    if parsed:
        return None
    # Marker sometimes appends visible prose to a URL destination after whitespace.
    destination = re.split(r"\s+", unescaped, maxsplit=1)[0]
    url = normalize_extracted_url(destination)
    if url:
        if is_doi_resolver_url(url):
            doi = doi_from_resolver_url(url)
            return ("doi", doi) if doi else None
        return ("url", url)
    doi = normalize_extracted_doi(destination)
    return ("doi", doi) if doi else None


def extract_text_identifiers_with_diagnostics(
    text: str,
) -> tuple[list[tuple[int, str, str]], list[dict[str, Any]]]:
    """Extract destination-first identifiers and rejected split DOI diagnostics."""

    candidates: list[tuple[int, int, str, str]] = []
    diagnostics: list[dict[str, Any]] = []
    links = scan_markdown_links(text)
    link_spans = [(link.start, link.end) for link in links]
    for link in links:
        classified = classify_text_destination(link.destination)
        if classified:
            candidates.append((link.start, 0, classified[0], classified[1]))
        else:
            parsed = parse_doi_destination(link.destination)
            if parsed and parsed.rejection_reason == "invalid_split_doi_continuation":
                diagnostics.append(
                    {
                        "position": link.start,
                        "reason": parsed.rejection_reason,
                        "candidate": parsed.raw_value,
                    }
                )
    masked = mask_spans(text, link_spans)
    autolink_spans: list[tuple[int, int]] = []
    for match in AUTOLINK_PATTERN.finditer(masked):
        classified = classify_text_destination(match.group(1))
        if classified:
            candidates.append((match.start(), 1, classified[0], classified[1]))
        autolink_spans.append(match.span())
    remaining = unescape_markdown_punctuation(mask_spans(masked, autolink_spans))
    resolver_spans: list[tuple[int, int]] = []
    for match in DOI_RESOLVER_PREFIX_PATTERN.finditer(remaining):
        parsed = parse_source_doi_candidate(remaining, match.end())
        resolver_spans.append((match.start(), max(parsed.end, match.end())))
        if parsed.value:
            candidates.append((match.start(), 2, "doi", parsed.value))
        elif parsed.rejection_reason == "invalid_split_doi_continuation":
            diagnostics.append(
                {
                    "position": match.start(),
                    "reason": parsed.rejection_reason or "invalid_doi_resolver_candidate",
                    "candidate": parsed.raw_value or remaining[match.start() : parsed.end],
                }
            )
    remaining_without_resolvers = mask_spans(remaining, resolver_spans)
    url_spans: list[tuple[int, int]] = []
    for match in PLAIN_URL_PATTERN.finditer(remaining_without_resolvers):
        classified = classify_text_destination(match.group(0))
        if classified:
            candidates.append((match.start(), 3, classified[0], classified[1]))
        url_spans.append(match.span())
    doi_text = mask_spans(remaining_without_resolvers, url_spans)
    cursor = 0
    while True:
        match = DOI_START_PATTERN.search(doi_text, cursor)
        if match is None:
            break
        parsed = parse_source_doi_candidate(doi_text, match.start())
        cursor = max(parsed.end, match.end(), match.start() + 1)
        if doi_text[parsed.end :].startswith("]("):
            wrapper_end = doi_text.find(")", parsed.end + 2)
            diagnostics.append(
                {
                    "position": match.start(),
                    "reason": "markdown_contamination",
                    "candidate": doi_text[
                        match.start() : wrapper_end + 1 if wrapper_end >= 0 else parsed.end + 2
                    ],
                }
            )
            continue
        if parsed.value:
            candidates.append((match.start(), 4, "doi", parsed.value))
        elif parsed.rejection_reason == "invalid_split_doi_continuation":
            diagnostics.append(
                {
                    "position": match.start(),
                    "reason": parsed.rejection_reason or "invalid_doi_candidate",
                    "candidate": parsed.raw_value,
                }
            )
    output: list[tuple[int, str, str]] = []
    seen: set[tuple[str, str]] = set()
    for position, stage, scheme, value in sorted(candidates, key=lambda item: (item[0], item[1], item[2], item[3])):
        del stage
        if (scheme, value) not in seen:
            seen.add((scheme, value))
            output.append((position, scheme, value))
    diagnostics.sort(key=lambda item: (item["position"], item["reason"], item["candidate"]))
    return output, diagnostics


def occurrence_level_doi_contradictions(
    text: str,
    identifiers: Sequence[tuple[int, str, str]],
    diagnostics: Sequence[Mapping[str, Any]],
) -> dict[tuple[int, str], Mapping[str, Any]]:
    """Map accepted DOI starts contradicted by a failed continuation in the same source construct."""

    unescaped = unescape_markdown_punctuation(text)
    contradictions: dict[tuple[int, str], Mapping[str, Any]] = {}
    for position, scheme, value in identifiers:
        if scheme != "doi":
            continue
        for diagnostic in diagnostics:
            diagnostic_position = int(diagnostic["position"])
            raw_candidate = unescape_markdown_punctuation(str(diagnostic["candidate"]))
            resolver = DOI_RESOLVER_PREFIX_PATTERN.match(raw_candidate)
            if resolver:
                raw_candidate = raw_candidate[resolver.end() :]
            if raw_candidate.casefold() == value.casefold() or not raw_candidate.casefold().startswith(
                value.casefold()
            ):
                continue
            same_start = diagnostic_position == position
            bridge = unescaped[position:diagnostic_position]
            markdown_label_destination = bool(
                re.fullmatch(
                    rf"(?:https?://(?:dx\.)?doi\.org/(?:doi:)?/?)?{re.escape(value)}\]\(\s*",
                    bridge,
                    re.IGNORECASE,
                )
            )
            if same_start or markdown_label_destination:
                contradictions[(position, value)] = diagnostic
                break
    return contradictions


def make_structural_deferred_doi_warning(
    context: str,
    candidate: str,
    evidence_text: str,
    source_location: Mapping[str, Any],
    diagnostic: Mapping[str, Any],
) -> dict[str, Any]:
    """Create an auditable deferral for one occurrence contradicted by its parser diagnostic."""

    return {
        "category": f"deferred_{context}_doi_candidate",
        "detail": {
            "action": "defer",
            "reason": f"same_source_occurrence_parser_contradiction:{diagnostic['reason']}",
            "context": context,
            "candidate": candidate,
            "evidence_text": evidence_text,
            "source_artifact": source_location["source_artifact"],
            "source_location": dict(source_location),
        },
    }


def extraction_warning_sort_key(warning: Mapping[str, Any]) -> tuple[int, str, str]:
    """Return one stable ordering key for malformed and deferred extraction warnings."""

    detail = warning["detail"]
    source_location = detail.get("source_location", {})
    line_number = int(detail.get("line_number", source_location.get("line_start", 0)))
    return line_number, str(detail["reason"]), str(detail["candidate"])


def extract_text_identifiers(text: str) -> list[tuple[int, str, str]]:
    """Extract destination-first DOI/URL identifiers from Markdown source text."""

    return extract_text_identifiers_with_diagnostics(text)[0]


def extract_dois_from_text(text: str) -> list[str]:
    """Extract strict DOI values from one Markdown source line."""

    return [value for _, scheme, value in extract_text_identifiers(text) if scheme == "doi"]


def is_malformed_doi_fragment(doi: str, source_text: str) -> bool:
    """Return whether local split evidence disproves an emitted DOI fragment."""

    normalized = normalize_extracted_doi(doi)
    if normalized != doi:
        return True
    identifiers, diagnostics = extract_text_identifiers_with_diagnostics(source_text)
    if any(
        value == doi and (position, value) in occurrence_level_doi_contradictions(
            source_text, identifiers, diagnostics
        )
        for position, scheme, value in identifiers
        if scheme == "doi"
    ):
        return True
    extracted = {value for _, scheme, value in identifiers if scheme == "doi"}
    if doi in extracted:
        return False
    return bool(diagnostics or any(value.startswith(doi) and value != doi for value in extracted))


def _continues_split_doi(line: str, following_line: str) -> bool:
    """Return whether a non-boundary next line can continue a DOI at line end."""

    if not following_line.strip() or HEADING_PATTERN.match(following_line):
        return False
    if re.match(r"^\s*(?:[-+*]\s+|\d+[.)]\s+)", following_line):
        return False
    match = list(DOI_START_PATTERN.finditer(unescape_markdown_punctuation(line)))
    if not match:
        return False
    tail = line[match[-1].start() :]
    return bool(re.search(r"10\.\d{4,9}/\S+\s*$", tail, re.IGNORECASE))


def extract_reference_dois_with_warnings(
    markdown: str,
    headings: Sequence[Mapping[str, Any]],
    canonical_artifact: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Extract reference DOIs and auditable malformed-candidate warnings."""

    lines = markdown.splitlines()
    records: dict[str, dict[str, Any]] = {}
    warnings: list[dict[str, Any]] = []
    seen_occurrences: set[tuple[str, int, int]] = set()
    seen_warnings: set[tuple[int, str, str]] = set()
    for index, heading in enumerate(headings):
        if heading["normalized_text"] not in REFERENCE_HEADINGS:
            continue
        start, end = reference_section_bounds(headings, index, len(lines))
        for line_number in range(start, end + 1):
            line = lines[line_number - 1]
            line_end = line_number
            scan_text = line
            if line_number < end and _continues_split_doi(line, lines[line_number]):
                scan_text = f"{line}\n{lines[line_number]}"
                line_end += 1
            identifiers, diagnostics = extract_text_identifiers_with_diagnostics(scan_text)
            contradictions = occurrence_level_doi_contradictions(scan_text, identifiers, diagnostics)
            for diagnostic in diagnostics:
                key = (line_number, diagnostic["reason"], diagnostic["candidate"])
                if key in seen_warnings:
                    continue
                seen_warnings.add(key)
                warnings.append(
                    make_extraction_warning(
                        "malformed_reference_doi_candidate",
                        diagnostic["reason"],
                        line_number,
                        diagnostic["candidate"],
                    )
                )
            for position, scheme, doi in identifiers:
                if scheme != "doi" or position >= len(line) + (1 if line_end > line_number else 0):
                    continue
                source_location = {
                    "source_artifact": canonical_artifact,
                    "section": heading["text"],
                    "line_start": line_number,
                    "line_end": line_end,
                }
                contradiction = contradictions.get((position, doi))
                if contradiction is not None:
                    warning = make_structural_deferred_doi_warning(
                        "reference",
                        doi,
                        scan_text.strip(),
                        source_location,
                        contradiction,
                    )
                    warning_key = (
                        line_number,
                        warning["detail"]["reason"],
                        warning["detail"]["candidate"],
                    )
                    if warning_key not in seen_warnings:
                        seen_warnings.add(warning_key)
                        warnings.append(warning)
                    continue
                occurrence_key = (doi, line_number, line_end)
                if occurrence_key in seen_occurrences:
                    continue
                seen_occurrences.add(occurrence_key)
                occurrence = {
                    "reference_text": scan_text.strip(),
                    "source_location": source_location,
                }
                if doi not in records:
                    records[doi] = {
                        "doi": doi,
                        "uri": f"https://doi.org/{doi}",
                        "reference_text": scan_text.strip(),
                        "source_location": occurrence["source_location"],
                        "occurrences": [],
                    }
                records[doi]["occurrences"].append(occurrence)
    warnings.sort(key=extraction_warning_sort_key)
    return list(records.values()), warnings


def extract_reference_dois(
    markdown: str,
    headings: Sequence[Mapping[str, Any]],
    canonical_artifact: str,
) -> list[dict[str, Any]]:
    """Extract DOI records only from mechanically identified reference sections."""

    return extract_reference_dois_with_warnings(markdown, headings, canonical_artifact)[0]


def reference_citation_fingerprint(value: str) -> str:
    """Create an exact presentation-normalized fingerprint for local citation reuse."""

    text = re.sub(r"<span[^>]*></span>", " ", value, flags=re.IGNORECASE)
    text = re.sub(r"!?\[[^]]*\]\([^)]*\)", " ", text)
    text = re.sub(r"<https?://[^>]+>", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"https?://\S+", " ", text, flags=re.IGNORECASE)
    text = unicodedata.normalize("NFKC", text).casefold()
    text = re.sub(r"[‐‑‒–—−]", "-", text)
    return " ".join(re.sub(r"[^a-z0-9]+", " ", text).split())


def _local_doi_extensions(doi: str, local_values: Iterable[str]) -> list[str]:
    """Return syntactically valid local DOI values that extend a candidate exactly."""

    extensions = []
    for value in local_values:
        if len(value) <= len(doi) or not value.startswith(doi):
            continue
        continuation = value[len(doi) :]
        if continuation[0] in ".-/_:;" and normalize_extracted_doi(value) == value:
            extensions.append(value)
    return sorted(set(extensions))


def reference_citation_evidence_key(value: str) -> tuple[str, int, int] | None:
    """Return exact title, year, and first-page evidence from a structured citation."""

    visible = re.sub(r"<span[^>]*></span>", " ", value, flags=re.IGNORECASE)
    visible = re.sub(r"!?\[[^]]*\]\([^)]*\)", " ", visible)
    visible = re.sub(r"<https?://[^>]+>", " ", visible, flags=re.IGNORECASE)
    visible = re.sub(r"https?://\S+", " ", visible, flags=re.IGNORECASE)
    year = re.search(r"(?:^|[,(\s])(?:18|19|20)\d{2}[a-z]?(?:\))?\s*[.,]", visible, re.IGNORECASE)
    if year is None:
        return None
    remainder = visible[year.end() :].lstrip(' \t\"“”')
    title = re.split(r"[.!?](?:[\"”']?\s|$)", remainder, maxsplit=1)[0]
    normalized = reference_citation_fingerprint(title)
    page_ranges = re.findall(r"(?<!\d)(\d{1,5})\s*[‐‑‒–—−-]\s*\d{1,5}(?!\d)", remainder)
    if len(normalized.split()) < 4 or not page_ranges:
        return None
    printed_year = int(re.search(r"(?:18|19|20)\d{2}", year.group(0)).group(0))
    return normalized, printed_year, int(page_ranges[-1])


def exact_local_extension_repair(
    doi: str,
    reference_text: str,
    local_references: Sequence[Mapping[str, Any]],
) -> str | None:
    """Return one longer DOI supported by an exact local citation-title match."""

    evidence_key = reference_citation_evidence_key(reference_text)
    if evidence_key is None:
        return None
    local_values = [str(reference["doi"]) for reference in local_references]
    extensions = set(_local_doi_extensions(doi, local_values))
    if not extensions:
        return None
    matches = {
        str(reference["doi"])
        for reference in local_references
        if str(reference["doi"]) in extensions
        and reference_citation_evidence_key(str(reference["reference_text"])) == evidence_key
    }
    return next(iter(matches)) if len(matches) == 1 else None


def reconcile_local_reference_doi_fragments(publications: Sequence[dict[str, Any]]) -> None:
    """Repair DOI prefixes only when exact local citation evidence identifies one extension."""

    all_records = [
        (record, reference)
        for record in publications
        for reference in record["content"]["reference_dois"]
    ]
    local_references = [reference for _, reference in all_records]

    for record in publications:
        retained: list[dict[str, Any]] = []
        added_warnings: list[dict[str, Any]] = []
        for reference in record["content"]["reference_dois"]:
            doi = reference["doi"]
            corrected = exact_local_extension_repair(
                doi,
                reference["reference_text"],
                local_references,
            )
            if corrected is None:
                retained.append(reference)
                continue
            line_number = reference["source_location"]["line_start"]
            reference["doi"] = corrected
            reference["uri"] = f"https://doi.org/{corrected}"
            added_warnings.append(
                {
                    "category": "repaired_reference_doi_candidate",
                    "detail": {
                        "reason": "exact_local_citation_extension",
                        "line_number": line_number,
                        "candidate": doi,
                        "corrected_doi": corrected,
                    },
                }
            )
            retained.append(reference)
        merged: dict[str, dict[str, Any]] = {}
        for reference in retained:
            doi = reference["doi"]
            if doi not in merged:
                merged[doi] = reference
                continue
            existing_keys = {
                (
                    occurrence["source_location"]["line_start"],
                    occurrence["source_location"]["line_end"],
                )
                for occurrence in merged[doi]["occurrences"]
            }
            for occurrence in reference["occurrences"]:
                key = (
                    occurrence["source_location"]["line_start"],
                    occurrence["source_location"]["line_end"],
                )
                if key not in existing_keys:
                    merged[doi]["occurrences"].append(occurrence)
                    existing_keys.add(key)
            merged[doi]["occurrences"].sort(
                key=lambda item: (
                    item["source_location"]["line_start"],
                    item["source_location"]["line_end"],
                    item["reference_text"],
                )
            )
        record["content"]["reference_dois"] = list(merged.values())
        record["reconciliation"]["warnings"].extend(added_warnings)
        record["reconciliation"]["warnings"].sort(
            key=lambda item: (item["category"], str(item["detail"]))
        )

    for record in publications:
        deduplicated: dict[tuple[str, str], dict[str, Any]] = {}
        for identifier in record["content"]["availability_identifiers"]:
            key = (identifier["identifier_scheme"], identifier["identifier_value"])
            deduplicated.setdefault(key, identifier)
        record["content"]["availability_identifiers"] = list(deduplicated.values())


def extract_availability_identifiers_with_warnings(
    markdown: str,
    headings: Sequence[Mapping[str, Any]],
    canonical_artifact: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Extract availability identifiers with the shared strict DOI parser."""

    lines = markdown.splitlines()
    output: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    seen_warnings: set[tuple[int, str, str]] = set()
    for index, heading in enumerate(headings):
        category = AVAILABILITY_HEADINGS.get(str(heading["normalized_text"]))
        if not category:
            continue
        start, end = section_bounds(headings, index, len(lines))
        for line_number in range(start, end + 1):
            line = lines[line_number - 1]
            terminal_offset = terminal_section_offset(line)
            if terminal_offset == 0:
                break
            stop_after_line = terminal_offset is not None
            if terminal_offset is not None:
                line = line[:terminal_offset]
            line_end = line_number
            scan_text = line
            if line_number < end and _continues_split_doi(line, lines[line_number]):
                scan_text = f"{line}\n{lines[line_number]}"
                line_end += 1
            identifiers, diagnostics = extract_text_identifiers_with_diagnostics(scan_text)
            contradictions = occurrence_level_doi_contradictions(scan_text, identifiers, diagnostics)
            for diagnostic in diagnostics:
                key = (line_number, diagnostic["reason"], diagnostic["candidate"])
                if key in seen_warnings:
                    continue
                seen_warnings.add(key)
                warnings.append(
                    make_extraction_warning(
                        "malformed_availability_doi_candidate",
                        diagnostic["reason"],
                        line_number,
                        diagnostic["candidate"],
                    )
                )
            for position, scheme, value in identifiers:
                if position >= len(line) + (1 if line_end > line_number else 0):
                    continue
                source_location = {
                    "source_artifact": canonical_artifact,
                    "line_start": line_number,
                    "line_end": line_end,
                }
                contradiction = contradictions.get((position, value)) if scheme == "doi" else None
                if contradiction is not None:
                    warning = make_structural_deferred_doi_warning(
                        "availability",
                        value,
                        scan_text.strip(),
                        source_location,
                        contradiction,
                    )
                    warning_key = (
                        line_number,
                        warning["detail"]["reason"],
                        warning["detail"]["candidate"],
                    )
                    if warning_key not in seen_warnings:
                        seen_warnings.add(warning_key)
                        warnings.append(warning)
                    continue
                if (scheme, value) in seen:
                    continue
                seen.add((scheme, value))
                output.append(
                    {
                        "section_category": category,
                        "section_title": heading["text"],
                        "identifier_scheme": scheme,
                        "identifier_value": value,
                        "identifier_uri": f"https://doi.org/{value}" if scheme == "doi" else value,
                        "evidence_text": scan_text.strip(),
                        "source_location": source_location,
                    }
                )
            if stop_after_line:
                break
    warnings.sort(key=extraction_warning_sort_key)
    return output, warnings


def extract_availability_identifiers(
    markdown: str,
    headings: Sequence[Mapping[str, Any]],
    canonical_artifact: str,
) -> list[dict[str, Any]]:
    """Extract strict identifiers from controlled availability sections."""

    return extract_availability_identifiers_with_warnings(markdown, headings, canonical_artifact)[0]


def read_json_object(path: Path, required: bool = True) -> dict[str, Any] | None:
    """Read a JSON object, failing for required missing or malformed files."""

    if not path.is_file():
        if required:
            raise CorpusBuildError(f"Required JSON file is missing: {path}")
        return None
    try:
        value = json.loads(read_utf8(path))
    except json.JSONDecodeError as exc:
        raise CorpusBuildError(f"JSON file is invalid: {path}: {exc}") from exc
    if not isinstance(value, dict):
        raise CorpusBuildError(f"JSON file must contain an object: {path}")
    return value


def source_file_paths(raw_root: Path, local_id: str) -> tuple[dict[str, Path], dict[str, str]]:
    """Construct the seven fixed publication source paths."""

    base = raw_root / "markdowns" / local_id
    actual = {
        "pdf_path": raw_root / "pdfs" / f"{local_id}.pdf",
        "markdown_path": base / "markdown" / f"{local_id}_md.md",
        "markdown_meta_path": base / "markdown" / f"{local_id}_md_meta.json",
        "chunks_path": base / "chunks" / f"{local_id}_chunks.json",
        "chunks_meta_path": base / "chunks" / f"{local_id}_chunks_meta.json",
        "marker_json_path": base / "json" / f"{local_id}_json.json",
        "marker_json_meta_path": base / "json" / f"{local_id}_json_meta.json",
    }
    persisted = {key: stored_path(raw_root, value) for key, value in actual.items()}
    return actual, persisted


def parse_markdown_front_matter_doi(markdown: str) -> str | None:
    """Read an explicit DOI from YAML front matter without interpreting prose."""

    if not markdown.startswith("---\n"):
        return None
    end = markdown.find("\n---\n", 4)
    if end < 0:
        return None
    try:
        metadata = yaml.safe_load(markdown[4:end])
    except yaml.YAMLError:
        return None
    return normalize_doi(metadata.get("doi")) if isinstance(metadata, dict) else None


def bib_field(entry: Entry | None, *names: str) -> Any:
    """Return the first non-empty case-insensitive BibTeX field."""

    if entry is None:
        return None
    fields = {key.casefold(): clean_scalar(value) for key, value in entry.fields.items()}
    for name in names:
        if fields.get(name.casefold()) is not None:
            return fields[name.casefold()]
    return None


def normalize_year(value: Any) -> int | None:
    """Normalize a four-digit publication year."""

    value = clean_scalar(value)
    if value is None:
        return None
    match = re.fullmatch(r"\d{4}", str(value).strip())
    return int(match.group(0)) if match else None


def build_conflicts(excel: Mapping[str, Any], entry: Entry | None) -> list[dict[str, Any]]:
    """Record substantive Excel/BibTeX disagreements under Excel authority."""

    if entry is None:
        return []
    comparisons = {
        "title": (excel.get("title"), bib_field(entry, "title")),
        "year": (normalize_year(excel.get("year")), normalize_year(bib_field(entry, "year"))),
        "doi": (normalize_doi(excel.get("doi")), normalize_doi(bib_field(entry, "doi"))),
        "url": (normalize_url(excel.get("url")), normalize_url(bib_field(entry, "url"))),
        "venue": (excel.get("journal"), bib_field(entry, "journal", "booktitle")),
    }
    conflicts: list[dict[str, Any]] = []
    for field, (excel_value, bibtex_value) in comparisons.items():
        if excel_value is None or bibtex_value is None:
            continue
        if comparison_text(excel_value) != comparison_text(bibtex_value):
            conflicts.append(
                {
                    "field": field,
                    "excel_value": excel_value,
                    "bibtex_value": bibtex_value,
                    "resolution": "excel_authority",
                }
            )
    return sorted(conflicts, key=lambda item: (item["field"], str(item["excel_value"]), str(item["bibtex_value"])))


def choose_value(override: Mapping[str, Any], excel_value: Any, bibtex_value: Any) -> Any:
    """Apply override, Excel, BibTeX, then null precedence for one field."""

    if override:
        return clean_scalar(override)
    return clean_scalar(excel_value) if clean_scalar(excel_value) is not None else clean_scalar(bibtex_value)


def build_publication_record(
    raw_root: Path,
    excel: Mapping[str, Any],
    bibtex: ParsedBibtex,
    override: Mapping[str, Any] | None,
    used_bibtex_keys: set[str],
) -> dict[str, Any]:
    """Build one complete publication record from its authoritative sources."""

    local_id = str(excel["id"])
    action = override.get("action") if override else None
    metadata = override.get("metadata", {}) if override else {}
    zotero_original = clean_scalar(excel.get("ZoteroID"))
    entry: Entry | None = None
    bibtex_key: str | None = None
    match_method: str
    if override:
        source_key = clean_scalar(override.get("source_zotero_key"))
        if source_key:
            if source_key not in bibtex.entries:
                raise CorpusBuildError(f"Override source_zotero_key does not exist for {local_id}: {source_key}")
            bibtex_key = str(source_key)
            entry = bibtex.entries[bibtex_key]
            used_bibtex_keys.add(bibtex_key)
        match_method = "override_replacement" if action == "replace_bibliographic_record" else "override_non_zotero"
    else:
        if not zotero_original:
            raise CorpusBuildError(f"Publication {local_id} has no Zotero key and no override")
        original = str(zotero_original)
        if original in bibtex.entries:
            bibtex_key = original
            match_method = "exact"
        else:
            bibtex_key = repair_zotero_key(original, bibtex.entries)
            if not bibtex_key:
                raise CorpusBuildError(f"Publication {local_id} has an unmatched Zotero key: {original}")
            match_method = "reversible_encoding_repair"
        entry = bibtex.entries[bibtex_key]
        used_bibtex_keys.add(bibtex_key)

    actual_paths, persisted_paths = source_file_paths(raw_root, local_id)
    if not actual_paths["pdf_path"].is_file():
        raise CorpusBuildError(f"Required PDF is missing for publication {local_id}: {actual_paths['pdf_path']}")
    raw_markdown = read_utf8(actual_paths["markdown_path"])
    warnings: list[dict[str, Any]] = []
    control_warning = audit_control_characters(raw_markdown, persisted_paths["markdown_path"])
    if control_warning:
        warnings.append(control_warning)
    markdown = sanitize_control_characters(raw_markdown)
    markdown_meta = read_json_object(actual_paths["markdown_meta_path"], required=False)
    if markdown_meta is None:
        warnings.append({"category": "missing_auxiliary_marker_meta", "detail": persisted_paths["markdown_meta_path"]})
    read_json_object(actual_paths["chunks_path"])
    read_json_object(actual_paths["chunks_meta_path"])
    read_json_object(actual_paths["marker_json_path"])
    read_json_object(actual_paths["marker_json_meta_path"])

    title = choose_value(metadata.get("title"), excel.get("title"), bib_field(entry, "title"))
    year = normalize_year(choose_value(metadata.get("year"), excel.get("year"), bib_field(entry, "year")))
    venue = choose_value(metadata.get("venue"), excel.get("journal"), bib_field(entry, "journal", "booktitle"))
    doi = normalize_doi(choose_value(metadata.get("doi"), excel.get("doi"), bib_field(entry, "doi")))
    if doi is None:
        doi = parse_markdown_front_matter_doi(markdown)
    url = normalize_url(choose_value(metadata.get("url"), excel.get("url"), bib_field(entry, "url")))
    if not doi and not url:
        raise CorpusBuildError(f"Publication {local_id} lacks both DOI and URL")
    identifiers = ([make_identifier("doi", doi)] if doi else []) + ([make_identifier("url", url)] if url else [])
    canonical_identifier = identifiers[0]
    canonical_artifact = canonical_identifier["uri"]
    headings = extract_headings(markdown)
    markdown_abstract, abstract_source, abstract_rejection_reason = extract_abstract_with_disposition(
        markdown,
        headings,
        canonical_artifact,
    )
    abstract = markdown_abstract or clean_scalar(bib_field(entry, "abstract", "abstractnote"))
    if abstract is not None and abstract_source is None:
        abstract_source = {"source_type": "bibtex_explicit", "source_artifact": canonical_artifact}
    if abstract_rejection_reason is not None:
        abstract_heading = next(
            (item for item in headings if item["normalized_text"] == "abstract"),
            None,
        )
        warnings.append(
            make_extraction_warning(
                "markdown_abstract_rejected",
                abstract_rejection_reason,
                int(abstract_heading["line_number"]) if abstract_heading else None,
            )
        )
    keywords, keyword_warnings = extract_keywords(
        markdown,
        headings,
        canonical_artifact,
        clean_scalar(bib_field(entry, "keywords", "keyword")),
    )
    warnings.extend(keyword_warnings)
    reference_dois, reference_warnings = extract_reference_dois_with_warnings(
        markdown,
        headings,
        canonical_artifact,
    )
    warnings.extend(reference_warnings)
    availability_identifiers, availability_warnings = extract_availability_identifiers_with_warnings(
        markdown,
        headings,
        canonical_artifact,
    )
    warnings.extend(availability_warnings)
    if metadata.get("authors") is not None:
        authors = authors_from_override(metadata["authors"])
    elif entry is not None:
        authors = authors_from_entry(entry, bibtex.raw_author_values.get(bibtex_key or "", []))
    else:
        authors = []
    bibtex_entry_type = clean_scalar(metadata.get("bibtex_entry_type")) or (entry.type.lower() if entry else None)
    record_type = clean_scalar(metadata.get("record_type")) or ENTRY_TYPE_MAP.get(str(bibtex_entry_type).lower())
    if record_type is None:
        record_type = "other"
        warnings.append({"category": "unknown_bibtex_entry_type", "detail": str(bibtex_entry_type)})
    conflicts = [] if override else build_conflicts(excel, entry)
    warnings.extend(
        {"category": "excel_bibtex_conflict", "detail": conflict["field"]} for conflict in conflicts
    )
    correction_of = override.get("correction_of") if override else None
    if correction_of is not None:
        if not isinstance(correction_of, dict):
            raise CorpusBuildError(f"correction_of must be an identifier mapping for {local_id}")
        scheme = correction_of.get("scheme")
        value = normalize_doi(correction_of.get("value")) if scheme == "doi" else normalize_url(correction_of.get("value"))
        if not value:
            raise CorpusBuildError(f"Invalid correction_of target for {local_id}")
        correction_of = make_identifier(str(scheme), value)
    page_stats = markdown_meta.get("page_stats") if isinstance(markdown_meta, dict) else None
    toc = markdown_meta.get("table_of_contents") if isinstance(markdown_meta, dict) else None
    sanitized_toc = sanitize_marker_value(toc) if isinstance(toc, list) else []
    return {
        "local_paper_id": local_id,
        "canonical_artifact_id": canonical_artifact,
        "canonical_identifier": canonical_identifier,
        "identifiers": identifiers,
        "record_type": record_type,
        "curation_status": "curated",
        "bibliographic": {
            "title": title,
            "authors": authors,
            "year": year,
            "venue": venue,
            "volume": choose_value(metadata.get("volume"), None, bib_field(entry, "volume")),
            "issue": choose_value(metadata.get("issue"), None, bib_field(entry, "number", "issue")),
            "pages": choose_value(metadata.get("pages"), None, bib_field(entry, "pages")),
            "publisher": choose_value(metadata.get("publisher"), None, bib_field(entry, "publisher")),
            "language": choose_value(metadata.get("language"), None, bib_field(entry, "language")),
            "abstract": abstract,
            "abstract_source": abstract_source,
        },
        "content": {
            "headings": headings,
            "explicit_keywords": keywords,
            "reference_dois": reference_dois,
            "availability_identifiers": availability_identifiers,
        },
        "document_structure": {
            "page_count": len(page_stats) if isinstance(page_stats, list) else None,
            "table_of_contents": sanitized_toc,
        },
        "source_files": persisted_paths,
        "bibliographic_relations": {"correction_of": correction_of},
        "reconciliation": {
            "excel_matched": True,
            "zotero_key_original": zotero_original,
            "bibtex_key": bibtex_key,
            "bibtex_match_method": match_method,
            "bibtex_entry_type": bibtex_entry_type,
            "override_applied": bool(override),
            "override_action": action,
            "conflicts": conflicts,
            "warnings": sorted(warnings, key=lambda item: (item["category"], str(item["detail"]))),
        },
    }


def calculate_summary(corpus: Mapping[str, Any], excel_count: int, bibtex_count: int) -> dict[str, Any]:
    """Calculate every summary field directly from emitted records."""

    publications = corpus["publications"]
    types = Counter(record["record_type"] for record in publications)
    with_doi = sum(any(identifier["scheme"] == "doi" for identifier in record["identifiers"]) for record in publications)
    record_warnings = sum(len(record["reconciliation"]["warnings"]) for record in publications)
    return {
        "excel_record_count": excel_count,
        "bibtex_entry_count": bibtex_count,
        "publication_count": len(publications),
        "with_doi": with_doi,
        "without_doi": len(publications) - with_doi,
        "without_doi_but_with_url": sum(
            not any(item["scheme"] == "doi" for item in record["identifiers"])
            and any(item["scheme"] == "url" for item in record["identifiers"])
            for record in publications
        ),
        "by_record_type": dict(sorted(types.items())),
        "exact_bibtex_matches": sum(record["reconciliation"]["bibtex_match_method"] == "exact" for record in publications),
        "encoding_repair_matches": sum(
            record["reconciliation"]["bibtex_match_method"] == "reversible_encoding_repair" for record in publications
        ),
        "override_record_count": sum(record["reconciliation"]["override_applied"] for record in publications),
        "non_zotero_record_count": sum(
            record["reconciliation"]["bibtex_match_method"] == "override_non_zotero" for record in publications
        ),
        "known_exclusion_count": len(corpus["known_exclusions"]),
        "pdf_count": sum(bool(record["source_files"]["pdf_path"]) for record in publications),
        "markdown_count": sum(bool(record["source_files"]["markdown_path"]) for record in publications),
        "chunks_count": sum(bool(record["source_files"]["chunks_path"]) for record in publications),
        "papers_with_abstract": sum(record["bibliographic"]["abstract"] is not None for record in publications),
        "papers_with_explicit_keywords": sum(bool(record["content"]["explicit_keywords"]) for record in publications),
        "explicit_keyword_count": sum(len(record["content"]["explicit_keywords"]) for record in publications),
        "papers_with_reference_dois": sum(bool(record["content"]["reference_dois"]) for record in publications),
        "reference_doi_count": sum(len(record["content"]["reference_dois"]) for record in publications),
        "papers_with_availability_identifiers": sum(
            bool(record["content"]["availability_identifiers"]) for record in publications
        ),
        "availability_identifier_count": sum(
            len(record["content"]["availability_identifiers"]) for record in publications
        ),
        "conflict_count": sum(len(record["reconciliation"]["conflicts"]) for record in publications),
        "warning_count": len(corpus["warnings"]) + record_warnings,
    }


def _occurrence_matches_discriminator(
    evidence_text: str,
    source_location: Mapping[str, Any],
    discriminator: Mapping[str, Any] | None,
) -> bool:
    """Match an optional exact, source-relative occurrence discriminator."""

    if discriminator is None:
        return True
    comparisons = {
        "evidence_text": evidence_text,
        "section": source_location.get("section"),
        "line_start": source_location.get("line_start"),
        "line_end": source_location.get("line_end"),
    }
    return all(comparisons.get(key) == value for key, value in discriminator.items())


def apply_identifier_dispositions(
    publications: Sequence[dict[str, Any]],
    overrides: Mapping[str, Mapping[str, Any]],
) -> None:
    """Apply exact source-scoped identifier dispositions and preserve their evidence."""

    by_id = {record["local_paper_id"]: record for record in publications}
    consumed_occurrences: set[tuple[str, str, str, int, int]] = set()
    for local_id in sorted(overrides, key=natural_publication_key):
        record = by_id[local_id]
        dispositions = overrides[local_id].get("identifier_dispositions", [])
        for index, disposition in enumerate(dispositions, start=1):
            context = str(disposition["context"])
            candidate = str(disposition["candidate"])
            discriminator = disposition.get("occurrence")
            matches: list[tuple[dict[str, Any], dict[str, Any]]] = []
            if context == "reference":
                for reference in record["content"]["reference_dois"]:
                    if reference["doi"] != candidate:
                        continue
                    for occurrence in reference["occurrences"]:
                        if _occurrence_matches_discriminator(
                            occurrence["reference_text"],
                            occurrence["source_location"],
                            discriminator,
                        ):
                            matches.append((reference, occurrence))
            else:
                for identifier in record["content"]["availability_identifiers"]:
                    if (
                        identifier["identifier_scheme"] == "doi"
                        and identifier["identifier_value"] == candidate
                        and _occurrence_matches_discriminator(
                            identifier["evidence_text"],
                            identifier["source_location"],
                            discriminator,
                        )
                    ):
                        matches.append((identifier, identifier))
            if len(matches) != 1:
                raise CorpusBuildError(
                    f"Identifier disposition {index} for {local_id} matched {len(matches)} occurrences; expected exactly one"
                )
            container, occurrence = matches[0]
            evidence_text = (
                occurrence["reference_text"] if context == "reference" else occurrence["evidence_text"]
            )
            source_location = occurrence["source_location"]
            occurrence_key = (
                local_id,
                context,
                candidate,
                int(source_location["line_start"]),
                int(source_location["line_end"]),
            )
            if occurrence_key in consumed_occurrences:
                raise CorpusBuildError(
                    f"Identifier disposition {index} for {local_id} duplicates an already matched occurrence"
                )
            consumed_occurrences.add(occurrence_key)
            if context == "reference":
                container["occurrences"].remove(occurrence)
                if container["occurrences"]:
                    primary = container["occurrences"][0]
                    container["reference_text"] = primary["reference_text"]
                    container["source_location"] = primary["source_location"]
                else:
                    record["content"]["reference_dois"].remove(container)
            else:
                record["content"]["availability_identifiers"].remove(container)
            record["reconciliation"]["warnings"].append(
                {
                    "category": f"deferred_{context}_doi_candidate",
                    "detail": {
                        "action": "defer",
                        "reason": str(disposition["reason"]),
                        "context": context,
                        "candidate": candidate,
                        "evidence_text": evidence_text,
                        "source_artifact": source_location["source_artifact"],
                        "source_location": source_location,
                    },
                }
            )
        record["reconciliation"]["warnings"].sort(
            key=lambda item: (item["category"], str(item["detail"]))
        )


def build_corpus(raw_root: Path, overrides_path: Path) -> dict[str, Any]:
    """Build the complete corpus in memory without writing output."""

    raw_root = raw_root.resolve()
    excel = load_excel_roster(raw_root / EXCEL_FILE)
    bibtex = load_bibtex(raw_root / BIBTEX_FILE)
    overrides = load_overrides(overrides_path.resolve())
    roster_ids = {record["id"] for record in excel}
    unknown_overrides = sorted(set(overrides) - roster_ids, key=natural_publication_key)
    if unknown_overrides:
        raise CorpusBuildError(f"Overrides target unknown local paper IDs: {unknown_overrides}")
    used_bibtex_keys: set[str] = set()
    publications = []
    for record in excel:
        override = overrides.get(record["id"])
        bibliographic_override = override if override and override.get("action") else None
        publications.append(
            build_publication_record(
                raw_root,
                record,
                bibtex,
                bibliographic_override,
                used_bibtex_keys,
            )
        )
    publications.sort(key=lambda record: natural_publication_key(record["local_paper_id"]))
    apply_identifier_dispositions(publications, overrides)
    reconcile_local_reference_doi_fragments(publications)
    exclusions: list[dict[str, Any]] = []
    for local_id, override in overrides.items():
        if override.get("action") != "replace_bibliographic_record":
            continue
        source_key = clean_scalar(override.get("source_zotero_key"))
        replacement_doi = normalize_doi(override["metadata"].get("doi"))
        replacement_url = normalize_url(override["metadata"].get("url"))
        if not source_key or not (replacement_doi or replacement_url):
            raise CorpusBuildError(f"Replacement override {local_id} lacks its source key or canonical identity")
        exclusions.append(
            {
                "source_type": "bibtex_entry",
                "source_key": source_key,
                "reason": "superseded_by_peer_reviewed_final",
                "replacement_canonical_artifact_id": (
                    f"https://doi.org/{replacement_doi}" if replacement_doi else replacement_url
                ),
            }
        )
    excluded_keys = {item["source_key"] for item in exclusions}
    unexplained = sorted(set(bibtex.entries) - used_bibtex_keys - excluded_keys)
    if unexplained:
        raise CorpusBuildError(f"Unexplained unused BibTeX entries: {unexplained}")
    corpus: dict[str, Any] = {
        "schema_version": SCHEMA_VERSION,
        "phase_a_version": PHASE_A_VERSION,
        "source": {
            "artifact_type": "publication",
            "raw_root": stored_raw_root(raw_root),
            "corpus_cutoff": "2026-03",
            "selection_method": "manually_curated_zotero_roster",
        },
        "publications": publications,
        "known_exclusions": sorted(exclusions, key=lambda item: item["source_key"]),
        "warnings": [],
        "summary": {},
    }
    known_dirs = {record["local_paper_id"] for record in publications}
    markdown_root = raw_root / "markdowns"
    if markdown_root.is_dir():
        for path in sorted(markdown_root.iterdir(), key=lambda item: item.name):
            if path.is_dir() and not path.name.startswith(".") and path.name not in known_dirs:
                corpus["warnings"].append(
                    {"category": "extra_raw_directory", "detail": stored_path(raw_root, path)}
                )
    corpus["warnings"].sort(key=lambda item: (item["category"], item["detail"]))
    corpus["summary"] = calculate_summary(corpus, len(excel), len(bibtex.entries))
    return corpus


def contains_forbidden_hash_key(value: Any) -> bool:
    """Return whether a nested object stores a forbidden per-file hash key."""

    if isinstance(value, dict):
        return any("sha256" in str(key).casefold() or contains_forbidden_hash_key(item) for key, item in value.items())
    if isinstance(value, list):
        return any(contains_forbidden_hash_key(item) for item in value)
    return False


def has_valid_source_line_range(location: Any, line_count: int) -> bool:
    """Return whether a source-location mapping has a valid inclusive line range."""

    if not isinstance(location, dict):
        return False
    start = location.get("line_start")
    end = location.get("line_end")
    return (
        isinstance(start, int)
        and not isinstance(start, bool)
        and isinstance(end, int)
        and not isinstance(end, bool)
        and 1 <= start <= end <= line_count
    )


def validate_corpus(
    corpus: Mapping[str, Any],
    raw_root: Path,
    expected_record_count: int,
    validate_frozen_snapshot: bool = False,
) -> dict[str, Any]:
    """Validate output structure, reconciliation, source coverage, and frozen anchors."""

    issues: list[str] = []
    required_top = {"schema_version", "phase_a_version", "source", "publications", "known_exclusions", "warnings", "summary"}
    if set(corpus) != required_top:
        issues.append(f"top-level keys differ: {sorted(set(corpus) ^ required_top)}")
    if corpus.get("schema_version") != SCHEMA_VERSION or corpus.get("phase_a_version") != PHASE_A_VERSION:
        issues.append("unsupported output version")
    try:
        source_excel_count = len(load_excel_roster(raw_root.resolve() / EXCEL_FILE))
        source_bibtex_count = len(load_bibtex(raw_root.resolve() / BIBTEX_FILE).entries)
    except CorpusBuildError as exc:
        issues.append(str(exc))
        source_excel_count = -1
        source_bibtex_count = -1
    publications = corpus.get("publications")
    if not isinstance(publications, list):
        issues.append("publications must be an array")
        publications = []
    if len(publications) != expected_record_count:
        issues.append(f"publication count {len(publications)} != expected {expected_record_count}")
    ids = [record.get("local_paper_id") for record in publications if isinstance(record, dict)]
    if ids != sorted(ids, key=natural_publication_key):
        issues.append("publication order is not natural local-ID order")
    if len(ids) != len(set(ids)):
        issues.append("duplicate local paper IDs")
    canonical_ids: list[str] = []
    expected_source_keys = {
        "pdf_path", "markdown_path", "markdown_meta_path", "chunks_path", "chunks_meta_path", "marker_json_path", "marker_json_meta_path"
    }
    required_record_keys = {
        "local_paper_id", "canonical_artifact_id", "canonical_identifier", "identifiers", "record_type",
        "curation_status", "bibliographic", "content", "document_structure", "source_files",
        "bibliographic_relations", "reconciliation",
    }
    required_bibliographic_keys = {
        "title", "authors", "year", "venue", "volume", "issue", "pages", "publisher", "language", "abstract", "abstract_source"
    }
    required_content_keys = {"headings", "explicit_keywords", "reference_dois", "availability_identifiers"}
    required_author_keys = {
        "position", "display_name", "given_names", "family_name", "name_particles", "suffix", "literal_name", "raw_bibtex"
    }
    required_reconciliation_keys = {
        "excel_matched", "zotero_key_original", "bibtex_key", "bibtex_match_method", "bibtex_entry_type",
        "override_applied", "override_action", "conflicts", "warnings",
    }
    for record in publications:
        if not isinstance(record, dict):
            issues.append("publication record is not an object")
            continue
        local_id = str(record.get("local_paper_id"))
        if set(record) != required_record_keys:
            issues.append(f"{local_id}: publication keys differ")
        canonical = record.get("canonical_artifact_id")
        if not isinstance(canonical, str) or not canonical:
            issues.append(f"{local_id}: missing canonical artifact")
        else:
            canonical_ids.append(canonical)
        identifiers = record.get("identifiers")
        if not isinstance(identifiers, list) or not identifiers:
            issues.append(f"{local_id}: identifiers are empty")
        else:
            schemes = [item.get("scheme") for item in identifiers if isinstance(item, dict)]
            if any(scheme not in {"doi", "url"} for scheme in schemes):
                issues.append(f"{local_id}: invalid identifier scheme")
            if "doi" in schemes and schemes[0] != "doi":
                issues.append(f"{local_id}: DOI is not the first identifier")
            if record.get("canonical_identifier") != identifiers[0] or canonical != identifiers[0].get("uri"):
                issues.append(f"{local_id}: canonical identity is inconsistent")
        bibliography = record.get("bibliographic", {})
        if set(bibliography) != required_bibliographic_keys:
            issues.append(f"{local_id}: bibliographic keys differ")
        if not clean_scalar(bibliography.get("title")):
            issues.append(f"{local_id}: title is empty")
        if normalize_year(bibliography.get("year")) is None:
            issues.append(f"{local_id}: year is invalid")
        if not clean_scalar(bibliography.get("venue")):
            issues.append(f"{local_id}: venue is empty")
        if not bibliography.get("authors"):
            issues.append(f"{local_id}: authors are empty")
        for expected_position, author in enumerate(bibliography.get("authors", []), start=1):
            if not isinstance(author, dict) or set(author) != required_author_keys:
                issues.append(f"{local_id}: author {expected_position} keys differ")
                continue
            if author.get("position") != expected_position or not clean_scalar(author.get("display_name")) or not clean_scalar(author.get("raw_bibtex")):
                issues.append(f"{local_id}: author {expected_position} is invalid")
        if set(record.get("content", {})) != required_content_keys:
            issues.append(f"{local_id}: content keys differ")
        if set(record.get("reconciliation", {})) != required_reconciliation_keys:
            issues.append(f"{local_id}: reconciliation keys differ")
        source_files = record.get("source_files", {})
        if set(source_files) != expected_source_keys:
            issues.append(f"{local_id}: source_files keys differ")
        for key, value in source_files.items():
            if not isinstance(value, str) or not value or Path(value).is_absolute() or ".." in Path(value).parts:
                issues.append(f"{local_id}: invalid stored source path {key}={value!r}")
        expected_actual, expected_persisted = source_file_paths(raw_root.resolve(), local_id)
        if source_files != expected_persisted:
            issues.append(f"{local_id}: stored source paths do not match deterministic paths")
        for key, path in expected_actual.items():
            if key == "markdown_meta_path":
                continue
            if not path.is_file():
                issues.append(f"{local_id}: required source missing: {key}")
        try:
            raw_markdown_text = read_utf8(expected_actual["markdown_path"])
            markdown_text = sanitize_control_characters(raw_markdown_text)
            markdown_line_count = len(markdown_text.splitlines())
        except CorpusBuildError as exc:
            issues.append(str(exc))
            raw_markdown_text = ""
            markdown_text = ""
            markdown_line_count = 0
        expected_control_warning = audit_control_characters(
            raw_markdown_text,
            expected_persisted["markdown_path"],
        )
        actual_control_warnings = [
            item
            for item in record.get("reconciliation", {}).get("warnings", [])
            if isinstance(item, dict) and item.get("category") == "unexpected_control_characters"
        ]
        expected_control_warnings = [expected_control_warning] if expected_control_warning else []
        if actual_control_warnings != expected_control_warnings:
            issues.append(f"{local_id}: control-character warning does not match raw Markdown audit")
        abstract_source = bibliography.get("abstract_source")
        if isinstance(abstract_source, dict) and abstract_source.get("source_type") == "markdown_explicit":
            if not has_valid_source_line_range(abstract_source, markdown_line_count):
                issues.append(f"{local_id}: abstract source line range is invalid")
            abstract_value = bibliography.get("abstract")
            if not isinstance(abstract_value, str):
                issues.append(f"{local_id}: Markdown abstract is not text")
            else:
                rejection_reason = validate_abstract_candidate(abstract_value)
                if rejection_reason:
                    issues.append(f"{local_id}: Markdown abstract is contaminated: {rejection_reason}")
            expected_abstract, expected_abstract_source, expected_rejection = extract_abstract_with_disposition(
                markdown_text,
                extract_headings(markdown_text),
                str(canonical),
            )
            if expected_rejection or bibliography.get("abstract") != expected_abstract or abstract_source != expected_abstract_source:
                issues.append(f"{local_id}: Markdown abstract does not match its accepted logical block")
        record_warning_categories = {
            item.get("category")
            for item in record.get("reconciliation", {}).get("warnings", [])
            if isinstance(item, dict)
        }
        if "markdown_abstract_rejected" in record_warning_categories:
            if isinstance(abstract_source, dict) and abstract_source.get("source_type") == "markdown_explicit":
                issues.append(f"{local_id}: rejected Markdown abstract remained selected")
            _, _, expected_rejection = extract_abstract_with_disposition(
                markdown_text,
                extract_headings(markdown_text),
                str(canonical),
            )
            if expected_rejection is None:
                issues.append(f"{local_id}: Markdown abstract rejection warning has no rejected candidate")
        content = record.get("content", {})
        for warning in record.get("reconciliation", {}).get("warnings", []):
            category = warning.get("category") if isinstance(warning, dict) else None
            if category not in {
                "deferred_reference_doi_candidate",
                "deferred_availability_doi_candidate",
            }:
                continue
            detail = warning.get("detail")
            context = "reference" if category == "deferred_reference_doi_candidate" else "availability"
            required_detail = {
                "action",
                "reason",
                "context",
                "candidate",
                "evidence_text",
                "source_artifact",
                "source_location",
            }
            if not isinstance(detail, dict) or set(detail) != required_detail:
                issues.append(f"{local_id}: deferred DOI warning detail is incomplete")
                continue
            candidate = detail.get("candidate")
            location = detail.get("source_location")
            if (
                detail.get("action") != "defer"
                or detail.get("context") != context
                or normalize_extracted_doi(candidate) != candidate
                or not clean_scalar(detail.get("reason"))
                or not clean_scalar(detail.get("evidence_text"))
                or detail.get("source_artifact") != canonical
                or not has_valid_source_line_range(location, markdown_line_count)
            ):
                issues.append(f"{local_id}: deferred DOI warning metadata is invalid")
                continue
            if context == "reference":
                extracted, extraction_warnings = extract_reference_dois_with_warnings(
                    markdown_text,
                    extract_headings(markdown_text),
                    str(canonical),
                )
                source_occurrences = [
                    occurrence
                    for reference in extracted
                    if reference["doi"] == candidate
                    for occurrence in reference["occurrences"]
                ]
                source_occurrences.extend(
                    {
                        "reference_text": item["detail"]["evidence_text"],
                        "source_location": item["detail"]["source_location"],
                    }
                    for item in extraction_warnings
                    if item.get("category") == "deferred_reference_doi_candidate"
                    and item.get("detail", {}).get("candidate") == candidate
                )
                accepted_occurrences = [
                    occurrence
                    for reference in content.get("reference_dois", [])
                    if reference.get("doi") == candidate
                    for occurrence in reference.get("occurrences", [])
                ]
                evidence_key = "reference_text"
            else:
                source_occurrences, extraction_warnings = extract_availability_identifiers_with_warnings(
                    markdown_text,
                    extract_headings(markdown_text),
                    str(canonical),
                )
                source_occurrences = [
                    identifier
                    for identifier in source_occurrences
                    if identifier["identifier_scheme"] == "doi"
                    and identifier["identifier_value"] == candidate
                ]
                source_occurrences.extend(
                    {
                        "evidence_text": item["detail"]["evidence_text"],
                        "source_location": item["detail"]["source_location"],
                    }
                    for item in extraction_warnings
                    if item.get("category") == "deferred_availability_doi_candidate"
                    and item.get("detail", {}).get("candidate") == candidate
                )
                accepted_occurrences = [
                    identifier
                    for identifier in content.get("availability_identifiers", [])
                    if identifier.get("identifier_scheme") == "doi"
                    and identifier.get("identifier_value") == candidate
                ]
                evidence_key = "evidence_text"
            exact_source_matches = [
                occurrence
                for occurrence in source_occurrences
                if occurrence.get(evidence_key) == detail["evidence_text"]
                and occurrence.get("source_location") == location
            ]
            exact_accepted_matches = [
                occurrence
                for occurrence in accepted_occurrences
                if occurrence.get(evidence_key) == detail["evidence_text"]
                and occurrence.get("source_location") == location
            ]
            if len(exact_source_matches) != 1 or exact_accepted_matches:
                issues.append(f"{local_id}: deferred DOI warning does not reconcile with one omitted source occurrence")
        explicit_keywords = content.get("explicit_keywords", [])
        expected_markdown_keywords, _ = extract_keywords(
            markdown_text,
            extract_headings(markdown_text),
            str(canonical),
            None,
        )
        actual_markdown_keywords = [
            item for item in explicit_keywords
            if isinstance(item, dict) and item.get("source_type") == "markdown_explicit"
        ]
        if actual_markdown_keywords != expected_markdown_keywords:
            issues.append(f"{local_id}: Markdown keywords do not match the accepted declaration blocks")
        seen_keywords: set[str] = set()
        for keyword in explicit_keywords:
            if not isinstance(keyword, dict):
                issues.append(f"{local_id}: explicit keyword is not an object")
                continue
            value = keyword.get("raw_value")
            normalized = keyword.get("value")
            if not isinstance(value, str) or not value.strip() or normalized != value.casefold():
                issues.append(f"{local_id}: explicit keyword is malformed: {value!r}")
                continue
            if clean_keyword_value(value) != value:
                issues.append(f"{local_id}: explicit keyword retains presentation punctuation: {value!r}")
            if value.startswith(("—", "–")):
                issues.append(f"{local_id}: explicit keyword retains a leading dash marker: {value!r}")
            if value.endswith(".") and not re.fullmatch(r"(?:[A-Za-z]\.){2,}", value):
                issues.append(f"{local_id}: explicit keyword retains a trailing sentence period: {value!r}")
            rejection_reason = validate_keyword_candidate(value)
            if rejection_reason:
                issues.append(f"{local_id}: invalid explicit keyword {value!r}: {rejection_reason}")
            if is_ambiguous_keyword_declaration(value):
                issues.append(f"{local_id}: ambiguous undelimited keyword was emitted: {value!r}")
            if normalized in seen_keywords:
                issues.append(f"{local_id}: duplicate explicit keyword: {normalized}")
            seen_keywords.add(str(normalized))
            location = keyword.get("source_location")
            if keyword.get("source_type") == "markdown_explicit" and not has_valid_source_line_range(
                location,
                markdown_line_count,
            ):
                issues.append(f"{local_id}: keyword source line range is invalid: {value!r}")
        seen_reference_dois: set[str] = set()
        for reference in content.get("reference_dois", []):
            doi = reference.get("doi") if isinstance(reference, dict) else None
            if normalize_extracted_doi(doi) != doi:
                issues.append(f"{local_id}: malformed extracted reference DOI: {doi!r}")
                continue
            if is_malformed_doi_fragment(doi, str(reference.get("reference_text", ""))):
                issues.append(f"{local_id}: truncated reference DOI contradicts local source evidence: {doi!r}")
            if reference.get("uri") != f"https://doi.org/{doi}":
                issues.append(f"{local_id}: reference DOI URI is inconsistent: {doi}")
            if doi in seen_reference_dois:
                issues.append(f"{local_id}: duplicate reference DOI record: {doi}")
            seen_reference_dois.add(doi)
            if not has_valid_source_line_range(reference.get("source_location"), markdown_line_count):
                issues.append(f"{local_id}: reference DOI source line range is invalid: {doi}")
            if not clean_scalar(reference.get("reference_text")):
                issues.append(f"{local_id}: reference DOI evidence text is empty: {doi}")
            occurrence_keys: set[tuple[int, int]] = set()
            occurrences = reference.get("occurrences", [])
            if not isinstance(occurrences, list) or not occurrences:
                issues.append(f"{local_id}: reference DOI has no occurrences: {doi}")
                occurrences = []
            for occurrence in occurrences:
                location = occurrence.get("source_location") if isinstance(occurrence, dict) else None
                if not has_valid_source_line_range(location, markdown_line_count):
                    issues.append(f"{local_id}: reference DOI occurrence range is invalid: {doi}")
                    continue
                key = (location["line_start"], location["line_end"])
                if key in occurrence_keys:
                    issues.append(f"{local_id}: duplicate reference DOI occurrence: {doi} at line {key[0]}")
                occurrence_keys.add(key)
                if not clean_scalar(occurrence.get("reference_text")):
                    issues.append(f"{local_id}: reference DOI occurrence evidence is empty: {doi}")
            ordered_occurrences = sorted(
                occurrences,
                key=lambda item: (
                    item.get("source_location", {}).get("line_start", 0),
                    item.get("source_location", {}).get("line_end", 0),
                    str(item.get("reference_text", "")),
                ),
            )
            if occurrences != ordered_occurrences:
                issues.append(f"{local_id}: reference DOI occurrences are not deterministically ordered: {doi}")
            if occurrences and (
                reference.get("reference_text") != occurrences[0].get("reference_text")
                or reference.get("source_location") != occurrences[0].get("source_location")
            ):
                issues.append(f"{local_id}: reference DOI primary evidence does not match first occurrence: {doi}")
        seen_availability: set[tuple[str, str]] = set()
        for identifier in content.get("availability_identifiers", []):
            if not isinstance(identifier, dict):
                issues.append(f"{local_id}: availability identifier is not an object")
                continue
            scheme = identifier.get("identifier_scheme")
            value = identifier.get("identifier_value")
            uri = identifier.get("identifier_uri")
            if scheme == "doi":
                if normalize_extracted_doi(value) != value or uri != f"https://doi.org/{value}":
                    issues.append(f"{local_id}: malformed availability DOI: {value!r}")
                elif is_malformed_doi_fragment(str(value), str(identifier.get("evidence_text", ""))):
                    issues.append(f"{local_id}: availability DOI contradicts local source evidence: {value!r}")
            elif scheme == "url":
                if normalize_extracted_url(value) != value or uri != value or is_doi_resolver_url(str(value)):
                    issues.append(f"{local_id}: malformed availability URL: {value!r}")
            else:
                issues.append(f"{local_id}: unknown availability identifier scheme: {scheme!r}")
            key = (str(scheme), str(value))
            if key in seen_availability:
                issues.append(f"{local_id}: duplicate availability identifier: {scheme}:{value}")
            seen_availability.add(key)
            if not has_valid_source_line_range(identifier.get("source_location"), markdown_line_count):
                issues.append(f"{local_id}: availability source line range is invalid: {scheme}:{value}")
        correction = record.get("bibliographic_relations", {}).get("correction_of")
        if correction is not None:
            if correction.get("scheme") == "doi":
                if normalize_doi(correction.get("value")) != correction.get("value"):
                    issues.append(f"{local_id}: correction_of DOI is invalid")
            elif correction.get("scheme") == "url":
                if normalize_url(correction.get("value")) != correction.get("value"):
                    issues.append(f"{local_id}: correction_of URL is invalid")
            else:
                issues.append(f"{local_id}: correction_of scheme is invalid")
    if len(canonical_ids) != len(set(canonical_ids)):
        issues.append("duplicate canonical artifact IDs")
    for finding in find_forbidden_control_characters(corpus):
        issues.append(
            "forbidden control characters remain at "
            f"{finding['path']}: {finding['code_points']} ({finding['occurrence_count']} occurrences)"
        )
    if contains_forbidden_hash_key(corpus):
        issues.append("corpus contains a forbidden per-file SHA-256 field")
    recomputed = calculate_summary(
        corpus,
        source_excel_count,
        source_bibtex_count,
    )
    if corpus.get("summary") != recomputed:
        issues.append("summary does not reconcile with publications")
    if validate_frozen_snapshot:
        summary = corpus.get("summary", {})
        anchors = {
            "excel_record_count": 228,
            "bibtex_entry_count": 227,
            "publication_count": 228,
            "with_doi": 227,
            "without_doi": 1,
            "without_doi_but_with_url": 1,
            "override_record_count": 2,
            "known_exclusion_count": 1,
            "reference_doi_count": 8856,
        }
        for key, expected in anchors.items():
            if summary.get(key) != expected:
                issues.append(f"frozen {key}={summary.get(key)!r}, expected {expected!r}")
        frozen_reference_values = {
            reference["doi"]
            for record in publications
            for reference in record["content"]["reference_dois"]
        }
        frozen_reference_occurrences = sum(
            len(reference["occurrences"])
            for record in publications
            for reference in record["content"]["reference_dois"]
        )
        if len(frozen_reference_values) != 6720:
            issues.append(
                f"frozen distinct reference DOI count={len(frozen_reference_values)!r}, expected 6720"
            )
        if frozen_reference_occurrences != 8963:
            issues.append(
                f"frozen reference DOI occurrence count={frozen_reference_occurrences!r}, expected 8963"
            )
        by_id = {record["local_paper_id"]: record for record in publications}
        frozen_records = {
            "71": ("Do land models miss key soil hydrological processes controlling soil moisture memory?", "https://doi.org/10.5194/hess-29-547-2025"),
            "87": (None, "https://doi.org/10.5194/hess-26-3377-2022"),
            "93": ("EASYMORE: A Python package to streamline the remapping of variables for Earth System models", None),
            "207": ("Nature-based solutions as buffers against coastal compound flooding: Exploring potential framework for process-based modeling of hazard mitigation", None),
        }
        for local_id, (title, canonical) in frozen_records.items():
            if local_id not in by_id:
                issues.append(f"frozen publication {local_id} is missing")
                continue
            if title is not None and by_id[local_id]["bibliographic"]["title"] != title:
                issues.append(f"frozen publication {local_id} title differs")
            if canonical is not None and by_id[local_id]["canonical_artifact_id"] != canonical:
                issues.append(f"frozen publication {local_id} canonical artifact differs")
        corrigendum = by_id.get("87-corrigendum")
        if not corrigendum or corrigendum["canonical_artifact_id"] != "https://doi.org/10.5194/hess-26-3377-2022-corrigendum":
            issues.append("frozen corrigendum canonical artifact differs")
        elif (
            corrigendum["record_type"] != "corrigendum"
            or corrigendum["bibliographic"]["year"] != 2023
            or corrigendum["bibliographic_relations"]["correction_of"]["uri"]
            != "https://doi.org/10.5194/hess-26-3377-2022"
        ):
            issues.append("frozen corrigendum metadata differs")
        no_doi = [record for record in publications if not any(item["scheme"] == "doi" for item in record["identifiers"])]
        if len(no_doi) != 1 or no_doi[0]["local_paper_id"] != "109" or no_doi[0]["canonical_identifier"]["scheme"] != "url":
            issues.append("frozen DOI-less publication is not ID 109 with URL identity")
    return {"valid": not issues, "issues": issues}


def serialize_corpus(corpus: Mapping[str, Any]) -> bytes:
    """Serialize corpus JSON deterministically with exactly one final newline."""

    return (json.dumps(corpus, indent=2, ensure_ascii=False, sort_keys=True) + "\n").encode("utf-8")


def write_corpus(path: Path, corpus: Mapping[str, Any]) -> str:
    """Atomically replace the corpus with validated bytes and return its digest."""

    data = serialize_corpus(corpus)
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="wb",
            dir=path.parent,
            prefix=f".{path.name}.",
            delete=False,
        ) as handle:
            temporary_path = Path(handle.name)
            handle.write(data)
            handle.flush()
            os.fsync(handle.fileno())
        os.chmod(temporary_path, 0o644)
        temporary_path.replace(path)
    finally:
        if temporary_path is not None and temporary_path.exists():
            temporary_path.unlink()
    return hashlib.sha256(data).hexdigest()


def print_report(corpus: Mapping[str, Any], validation: Mapping[str, Any], output: Path, digest: str | None) -> None:
    """Print the concise validation report required by the Phase A contract."""

    summary = corpus["summary"]
    warning_counts: Counter[str] = Counter(item["category"] for item in corpus["warnings"])
    for record in corpus["publications"]:
        warning_counts.update(item["category"] for item in record["reconciliation"]["warnings"])
    print(f"schema version: {corpus['schema_version']}")
    print(f"Phase A version: {corpus['phase_a_version']}")
    print(f"Excel records: {summary['excel_record_count']}")
    print(f"BibTeX entries: {summary['bibtex_entry_count']}")
    print(f"publications emitted: {summary['publication_count']}")
    print(f"records by publication type: {json.dumps(summary['by_record_type'], sort_keys=True)}")
    print(f"publications with DOI: {summary['with_doi']}")
    print(f"publications without DOI: {summary['without_doi']}")
    print(f"exact BibTeX matches: {summary['exact_bibtex_matches']}")
    print(f"encoding-repair matches: {summary['encoding_repair_matches']}")
    print(f"override records: {summary['override_record_count']}")
    print(f"known exclusions: {summary['known_exclusion_count']}")
    print(f"PDFs found: {summary['pdf_count']}")
    print(f"Markdown files found: {summary['markdown_count']}")
    print(f"chunks files found: {summary['chunks_count']}")
    print(f"papers with explicit abstracts: {summary['papers_with_abstract']}")
    print(f"papers with explicit keywords: {summary['papers_with_explicit_keywords']}")
    print(f"reference DOIs: {summary['reference_doi_count']}")
    print(f"availability identifiers: {summary['availability_identifier_count']}")
    print(f"metadata conflicts: {summary['conflict_count']}")
    print(f"warnings by category: {json.dumps(dict(sorted(warning_counts.items())), sort_keys=True)}")
    print(f"validation status: {'valid' if validation['valid'] else 'invalid'}")
    for issue in validation["issues"]:
        print(f"validation issue: {issue}")
    print(f"output path: {output}")
    if digest:
        print(f"output SHA-256: {digest}")


def build_argument_parser() -> argparse.ArgumentParser:
    """Create the publication preprocessing command-line parser."""

    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--raw-root", type=Path, default=DEFAULT_RAW_ROOT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--overrides", type=Path, default=DEFAULT_OVERRIDES)
    parser.add_argument("--expected-record-count", type=int, default=DEFAULT_EXPECTED_RECORD_COUNT)
    parser.add_argument("--validate-frozen-snapshot", action="store_true")
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    """Build, validate, then write the deterministic publication corpus."""

    args = build_argument_parser().parse_args(argv)
    raw_root = resolve_path(args.raw_root)
    output = resolve_path(args.output)
    overrides = resolve_path(args.overrides)
    try:
        corpus = build_corpus(raw_root, overrides)
        validation = validate_corpus(corpus, raw_root, args.expected_record_count, args.validate_frozen_snapshot)
        if not validation["valid"]:
            print_report(corpus, validation, output, None)
            return 1
        digest = write_corpus(output, corpus)
        print_report(corpus, validation, output, digest)
        return 0
    except CorpusBuildError as exc:
        print(f"publication preprocessing failed: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
